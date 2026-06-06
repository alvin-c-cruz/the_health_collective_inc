import json
import os
from io import BytesIO

import openpyxl
from flask import (
    Blueprint,
    Response,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import current_user
from sqlalchemy.exc import IntegrityError
from werkzeug.utils import secure_filename

from application.blueprints.audit.utils import (
    log_create,
    log_delete,
    model_to_dict,
)
from application.blueprints.user import login_required, roles_accepted
from application.extensions import db

from . import app_label, app_name
from .forms import Form
from .models import Measure as Obj
from .models import ObjAdmin as Approver
from .models import ObjUser as Preparer

bp = Blueprint(app_name, __name__, template_folder="pages", url_prefix=f"/{app_name}")
ROLES_ACCEPTED = app_label


@bp.route("/")
@login_required
@roles_accepted([ROLES_ACCEPTED])
def home():
    rows = Obj.query.order_by(getattr(Obj, f"{app_name}_name")).all()

    context = {"rows": rows}

    return render_template(f"{app_name}/home.html", **context)


@bp.route("/add", methods=["POST", "GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def add():
    if request.method == "POST":
        form = Form()
        form._post(request.form, current_user.id)

        if form._validate_on_submit():
            form._save()
            return redirect(url_for(f"{app_name}.home"))
    else:
        form = Form()

    context = {
        "form": form,
    }

    return render_template(f"{app_name}/form.html", **context)


@bp.route("/edit/<int:record_id>", methods=["POST", "GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def edit(record_id):
    if request.method == "POST":
        form = Form()
        form._post(request.form, current_user.id)

        if form._validate_on_submit():
            form._save()
            return redirect(url_for(f"{app_name}.home"))

    else:
        obj = Obj.query.get(record_id)
        form = Form()
        form._populate(obj)

    context = {
        "form": form,
    }

    return render_template(f"{app_name}/form.html", **context)


@bp.route("/delete/<int:record_id>", methods=["POST", "GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def delete(record_id):
    obj = Obj.query.get_or_404(record_id)
    preparer = obj.preparer

    # Capture values before deletion
    old_values = model_to_dict(obj, ["measure_name"])
    record_id_for_log = obj.id
    identifier = str(obj)

    try:
        if preparer:
            db.session.delete(preparer)
        db.session.delete(obj)

        # Log deletion before commit
        log_delete(
            module="measure",
            record_id=record_id_for_log,
            record_identifier=identifier,
            old_values=old_values,
        )

        db.session.commit()
        flash(f"{identifier} has been deleted.", category="success")
    except IntegrityError:
        db.session.rollback()
        flash(f"Cannot delete {obj} because it has related records.", category="error")

    return redirect(url_for(f"{app_name}.home"))


@bp.route("/approve/<int:record_id>", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def approve(record_id):
    if not current_user.admin:
        flash("Administrator rights required.", category="error")
        return redirect(url_for(f"{app_name}.home"))

    obj = Obj.query.get_or_404(record_id)

    data = {f"{app_name}_id": record_id, "user_id": current_user.id}

    approve = Approver(**data)

    db.session.add(approve)
    db.session.commit()

    flash(f"Approved: {getattr(obj, f'{app_name}_name')}", category="success")
    return redirect(url_for(f"{app_name}.home"))


@bp.route("/unlock/<int:record_id>", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def unlock(record_id):
    if not current_user.admin:
        flash("Administrator rights required.", category="error")
        return redirect(url_for(f"{app_name}.home"))

    obj = Obj.query.get_or_404(record_id)

    data = {
        f"{app_name}_id": record_id,
    }

    approve = Approver.query.filter_by(**data).first()

    db.session.delete(approve)
    db.session.commit()

    flash(f"Unlocked: {getattr(obj, f'{app_name}_name')}", category="error")
    return redirect(url_for(f"{app_name}.home"))


@bp.route("/autocomplete", methods=["GET"])
@login_required
def _autocomplete():
    options = [
        getattr(i, f"{app_name}_name")
        for i in Obj.query.order_by(getattr(Obj, f"{app_name}_name")).all()
    ]
    return Response(json.dumps(options), mimetype="application/json")


@bp.route("/upload", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def upload():
    file = request.files.get("xlsx_file")

    temp_dir = os.path.join(current_app.instance_path, "temp")
    os.makedirs(temp_dir, exist_ok=True)

    # Clean old files
    for f in os.listdir(temp_dir):
        os.remove(os.path.join(temp_dir, f))

    if not file or not file.filename.endswith(".xlsx"):
        flash("Please upload a valid .xlsx file.", "danger")
        return redirect(url_for(f"{app_name}.home"))

    filepath = os.path.join(temp_dir, secure_filename(file.filename))
    file.save(filepath)

    try:
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active

        imported = 0
        skipped = 0

        title = sheet.title
        cell_measure_name = sheet["A1"].value

        checker = (title, cell_measure_name)

        if checker == ("Measures", "Measure"):
            for row in sheet.iter_rows(min_row=2, values_only=True):
                measure_name = row[0]

                if not measure_name:
                    continue

                existing = Obj.query.filter(
                    Obj.measure_name == str(measure_name)
                ).first()

                if existing:
                    skipped += 1
                    continue

                measure = Obj(measure_name=str(measure_name).upper())

                db.session.add(measure)
                db.session.flush()

                # Log creation after flush to get ID
                log_create(
                    module="measure",
                    record_id=measure.id,
                    record_identifier=str(measure),
                    new_values=model_to_dict(measure, ["measure_name"]),
                    notes="Measure imported from Excel",
                )

                db.session.commit()

                preparer_data = {
                    f"{app_name}_id": measure.id,
                    "user_id": current_user.id,
                }
                preparer = Preparer(**preparer_data)
                db.session.add(preparer)
                db.session.commit()

                imported += 1
            flash(
                f"{imported} record(s) imported successfully. {skipped} skipped due to duplicates.",
                "success",
            )
        else:
            flash("Error processing file: Invalid format.", "danger")

    except Exception as e:
        flash(f"Error processing file: {e!s}", "danger")

    return redirect(url_for(f"{app_name}.home"))


@bp.route("/download-template", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Measures"

    # Header row
    ws.append(["Measure"])

    # Save workbook to memory
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="measure.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
