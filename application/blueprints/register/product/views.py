from flask import Blueprint, render_template, render_template_string, request, redirect, url_for, flash, Response, send_file, current_app
import os
import json
import openpyxl
from io import BytesIO
from werkzeug.utils import secure_filename
from sqlalchemy.exc import IntegrityError
from .models import Product as Obj
from .models import ObjAdmin as Approver
from .models import ObjUser as Preparer
from .forms import Form
from application.extensions import db
from application.blueprints.user import login_required, roles_accepted
from flask_login import current_user

from . import app_name, app_label


bp = Blueprint(app_name, __name__, template_folder="pages", url_prefix=f"/{app_name}")
ROLES_ACCEPTED = app_label


@bp.route("/")
@login_required
@roles_accepted([ROLES_ACCEPTED])
def home():
    rows = Obj.query.order_by(getattr(Obj, f"{app_name}_name")).all()

    context = {
        "rows": rows
    }

    return render_template(f"{app_name}/home.html", **context)


@bp.route("/add", methods=["POST", "GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def add():
    popup = request.args.get('popup') == '1'
    if request.method == "POST":
        form = Form()
        form._post(request.form, current_user.id)

        if form._validate_on_submit():
            form._save()
            if popup:
                new_obj = Obj.query.filter_by(product_name=form.product_name).order_by(Obj.id.desc()).first()
                return render_template_string(
                    '<!doctype html><html><head><meta charset="utf-8"></head><body>'
                    '<script>'
                    'if(window.opener){'
                    'window.opener.postMessage({type:"product_added",product_id:{{ pid }},product_name:{{ pname | tojson }},product_type_name:{{ ptype | tojson }}},"*");'
                    '}'
                    'window.close();'
                    '</script>'
                    '<p style="font-family:sans-serif;padding:2rem;">Product saved. This window will close automatically.</p>'
                    '</body></html>',
                    pid=new_obj.id if new_obj else 0,
                    pname=form.product_name,
                    ptype=form.product_type_name
                )
            return redirect(url_for(f'{app_name}.home'))
    else:
        form = Form()

    context = {
        "form": form,
        "popup": popup,
    }

    return render_template(f"{app_name}/form.html", **context)


@bp.route(f"/edit/<int:record_id>", methods=["POST", "GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def edit(record_id):   
    if request.method == "POST":
        form = Form()
        form._post(request.form, current_user.id)

        if form._validate_on_submit():
            form._save()
            return redirect(url_for(f'{app_name}.home'))

    else:
        obj = Obj.query.get(record_id)
        form = Form()
        form._populate(obj)

    context = {
        "form": form,
    }

    return render_template(f"{app_name}/form.html", **context)


@bp.route("/delete/<int:record_id>", methods=["POST", "GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def delete(record_id):
    obj = Obj.query.get_or_404(record_id)
    preparer = obj.preparer

    try:
        # Check for related transaction details
        from application.blueprints.operations.daily_sales.models import TransactionDetail
        related_details = TransactionDetail.query.filter_by(product_id=record_id).count()

        if related_details > 0:
            flash(f"Cannot delete {obj} because it is used in {related_details} transaction(s).", category="error")
            return redirect(url_for(f'{app_name}.home'))

        # Delete related admin/user records first
        approver = obj.approved
        if approver:
            db.session.delete(approver)
        if preparer:
            db.session.delete(preparer)

        # Now delete the product
        db.session.delete(obj)
        db.session.commit()
        flash(f"{obj} has been deleted.", category="success")

    except (IntegrityError, AssertionError) as e:
        db.session.rollback()
        flash(f"Cannot delete {obj} because it has related records.", category="error")
    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting {obj}: {str(e)}", category="error")

    return redirect(url_for(f'{app_name}.home'))


@bp.route("/approve/<int:record_id>", methods=['GET'])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def approve(record_id):
    if not current_user.admin:
        flash("Administrator rights required.", category="error")
        return redirect(url_for(f"{app_name}.home"))
    
    obj = Obj.query.get_or_404(record_id)

    data = {
        f"{app_name}_id": record_id,
        "user_id": current_user.id
    }

    approve = Approver(**data)

    db.session.add(approve)
    db.session.commit()

    flash(f"Approved: {getattr(obj, f"{app_name}_name")}", category="success")
    return redirect(url_for(f"{app_name}.home"))   
    

@bp.route("/unlock/<int:record_id>", methods=['GET'])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def unlock(record_id):
    if not current_user.admin:
        flash("Administrator rights required.", category="error")
        return redirect(url_for(f"{app_name}.home"))
    
    obj = Obj.query.get_or_404(record_id)

    data = {
        f"{app_name}_id": record_id,
    }

    approve = Approver.query.filter_by(**data).first()
    
    db.session.delete(approve)
    db.session.commit()

    flash(f"Unlocked: {getattr(obj, f"{app_name}_name")}", category="error")
    return redirect(url_for(f"{app_name}.home"))   
    

@bp.route("/autocomplete", methods=['GET'])
@login_required
def _autocomplete():
    options = [getattr(i,f"{app_name}_name") for i in Obj.query.order_by(getattr(Obj,f"{app_name}_name")).all()]
    return Response(json.dumps(options), mimetype='application/json')


@bp.route("/upload", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def upload():
    file = request.files.get("xlsx_file")

    temp_dir = os.path.join(current_app.instance_path, "temp")
    os.makedirs(temp_dir, exist_ok=True)

    # Clean old files
    for f in os.listdir(temp_dir):
        os.remove(os.path.join(temp_dir, f))

    if not file or not file.filename.endswith(".xlsx"):
        flash("Please upload a valid .xlsx file.", "danger")
        return redirect(url_for(f"{app_name}.home"))

    filepath = os.path.join(temp_dir, secure_filename(file.filename))
    file.save(filepath)

    try:
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active

        imported = 0
        skipped = 0
        
        title = sheet.title
        cell_product_name = sheet["A1"].value
              
        checker = (title, cell_product_name)
    
        if checker == ("Products", "Product Name"):
            for row in sheet.iter_rows(min_row=2, values_only=True):
                product_name = row[0]

                if not product_name:
                    continue

                existing = Obj.query.filter(
                    (Obj.product_name == str(product_name)) 
                ).first()

                if existing:
                    skipped += 1
                    continue

                product = Obj(
                    product_name=str(product_name).upper()
                )

                db.session.add(product)
                db.session.commit()

                preparer_data = {
                    f"{app_name}_id": product.id,
                    "user_id": current_user.id
                }
                preparer = Preparer(**preparer_data)
                db.session.add(preparer)
                db.session.commit()

                imported += 1
            flash(f"{imported} record(s) imported successfully. {skipped} skipped due to duplicates.", "success")
        else:
            flash(f"Error processing file: Invalid format.", "danger")

    except Exception as e:
        flash(f"Error processing file: {str(e)}", "danger")

    return redirect(url_for(f"{app_name}.home"))


@bp.route("/download-template", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    # Header row
    ws.append(["Product Name"])

    # Save workbook to memory
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="product.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
