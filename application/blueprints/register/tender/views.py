import json
import os
from io import BytesIO

import openpyxl
from flask import (
    Blueprint,
    Response,
    current_app,
    flash,
    redirect,
    render_template,
    render_template_string,
    request,
    send_file,
    url_for,
)
from flask_login import current_user
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from werkzeug.utils import secure_filename

from application.blueprints.audit.utils import (
    log_create,
    log_delete,
    model_to_dict,
)
from application.blueprints.user import login_required, roles_accepted
from application.extensions import db

from . import app_label, app_name
from .forms import Form
from .models import ObjAdmin as Approver
from .models import ObjUser as Preparer
from .models import Tender as Obj

bp = Blueprint(app_name, __name__, template_folder="pages", url_prefix=f"/{app_name}")
ROLES_ACCEPTED = app_label


@bp.route("/")
@login_required
@roles_accepted([ROLES_ACCEPTED])
def home():
    rows = Obj.query.order_by(Obj.sort_order.desc()).all()

    context = {"rows": rows}

    return render_template(f"{app_name}/home.html", **context)


@bp.route("/add", methods=["POST", "GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def add():
    popup = request.args.get("popup") == "1"
    if request.method == "POST":
        form = Form()
        form._post(request.form, current_user.id)
        if not form.sort_order:
            max_order = db.session.query(db.func.max(Obj.sort_order)).scalar() or 0
            form.sort_order = max_order + 10

        if form._validate_on_submit():
            form._save()
            if popup:
                new_obj = (
                    Obj.query.filter_by(tender_name=form.tender_name)
                    .order_by(Obj.id.desc())
                    .first()
                )
                return render_template_string(
                    '<!doctype html><html><head><meta charset="utf-8"></head><body>'
                    "<script>"
                    "if(window.opener){"
                    'window.opener.postMessage({type:"tender_added",tender_id:{{ tid }},tender_name:{{ tname | tojson }}},"*");'
                    "}"
                    "window.close();"
                    "</script>"
                    '<p style="font-family:sans-serif;padding:2rem;">Tender saved. This window will close automatically.</p>'
                    "</body></html>",
                    tid=new_obj.id if new_obj else 0,
                    tname=form.tender_name,
                )
            return redirect(url_for(f"{app_name}.home"))
    else:
        form = Form()

    context = {
        "form": form,
        "popup": popup,
    }

    return render_template(f"{app_name}/form.html", **context)


@bp.route("/edit/<int:record_id>", methods=["POST", "GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def edit(record_id):
    if request.method == "POST":
        form = Form()
        form._post(request.form, current_user.id)

        if form._validate_on_submit():
            form._save()
            return redirect(url_for(f"{app_name}.home"))

    else:
        obj = Obj.query.get(record_id)
        form = Form()
        form._populate(obj)

    context = {
        "form": form,
    }

    return render_template(f"{app_name}/form.html", **context)


@bp.route("/delete/<int:record_id>", methods=["POST", "GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def delete(record_id):
    obj = Obj.query.get_or_404(record_id)
    preparer = obj.preparer

    # Capture values before deletion
    old_values = model_to_dict(
        obj,
        [
            "tender_name",
            "symbol",
            "transaction_types",
            "sort_order",
            "report_static",
            "is_receivable",
        ],
    )
    record_id_for_log = obj.id
    identifier = str(obj)

    try:
        if preparer:
            db.session.delete(preparer)
        db.session.delete(obj)

        # Log deletion before commit
        log_delete(
            module="tender",
            record_id=record_id_for_log,
            record_identifier=identifier,
            old_values=old_values,
        )

        db.session.commit()
        flash(f"{identifier} has been deleted.", category="success")
    except IntegrityError:
        db.session.rollback()
        flash(f"Cannot delete {obj} because it has related records.", category="error")

    return redirect(url_for(f"{app_name}.home"))


@bp.route("/approve/<int:record_id>", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def approve(record_id):
    if not current_user.admin:
        flash("Administrator rights required.", category="error")
        return redirect(url_for(f"{app_name}.home"))

    obj = Obj.query.get_or_404(record_id)

    data = {f"{app_name}_id": record_id, "user_id": current_user.id}

    approve = Approver(**data)

    db.session.add(approve)
    db.session.commit()

    flash(f"Approved: {getattr(obj, f'{app_name}_name')}", category="success")
    return redirect(url_for(f"{app_name}.home"))


@bp.route("/unlock/<int:record_id>", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def unlock(record_id):
    if not current_user.admin:
        flash("Administrator rights required.", category="error")
        return redirect(url_for(f"{app_name}.home"))

    obj = Obj.query.get_or_404(record_id)

    data = {
        f"{app_name}_id": record_id,
    }

    approve = Approver.query.filter_by(**data).first()

    db.session.delete(approve)
    db.session.commit()

    flash(f"Unlocked: {getattr(obj, f'{app_name}_name')}", category="error")
    return redirect(url_for(f"{app_name}.home"))


@bp.route("/autocomplete", methods=["GET"])
@login_required
def _autocomplete():
    options = [
        getattr(i, f"{app_name}_name")
        for i in Obj.query.order_by(getattr(Obj, f"{app_name}_name")).all()
    ]
    return Response(json.dumps(options), mimetype="application/json")


@bp.route("/upload", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def upload():
    file = request.files.get("xlsx_file")

    temp_dir = os.path.join(current_app.instance_path, "temp")
    os.makedirs(temp_dir, exist_ok=True)

    # Clean old files
    for f in os.listdir(temp_dir):
        os.remove(os.path.join(temp_dir, f))

    if not file or not file.filename.endswith(".xlsx"):
        flash("Please upload a valid .xlsx file.", "danger")
        return redirect(url_for(f"{app_name}.home"))

    filepath = os.path.join(temp_dir, secure_filename(file.filename))
    file.save(filepath)

    try:
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active

        imported = 0
        skipped = 0

        title = sheet.title
        cell_product_name = sheet["A1"].value

        checker = (title, cell_product_name)

        if checker == ("Tenders", "Tender Name"):
            for row in sheet.iter_rows(min_row=2, values_only=True):
                tender_name = row[0]

                if not tender_name:
                    continue

                # Case-insensitive duplicate check
                existing = Obj.query.filter(
                    func.lower(Obj.tender_name) == func.lower(str(tender_name))
                ).first()

                if existing:
                    skipped += 1
                    continue

                tender = Obj(tender_name=str(tender_name).upper())

                db.session.add(tender)
                db.session.flush()

                # Log creation after flush to get ID
                log_create(
                    module="tender",
                    record_id=tender.id,
                    record_identifier=str(tender),
                    new_values=model_to_dict(
                        tender,
                        [
                            "tender_name",
                            "symbol",
                            "transaction_types",
                            "sort_order",
                            "report_static",
                            "is_receivable",
                        ],
                    ),
                    notes="Tender imported from Excel",
                )

                db.session.commit()

                preparer_data = {
                    f"{app_name}_id": tender.id,
                    "user_id": current_user.id,
                }
                preparer = Preparer(**preparer_data)
                db.session.add(preparer)
                db.session.commit()

                imported += 1
            flash(
                f"{imported} record(s) imported successfully. {skipped} skipped due to duplicates.",
                "success",
            )
        else:
            flash("Error processing file: Invalid format.", "danger")

    except Exception as e:
        flash(f"Error processing file: {e!s}", "danger")

    return redirect(url_for(f"{app_name}.home"))


@bp.route("/download-template", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tenders"

    # Header row
    ws.append(["Tender Name"])

    # Save workbook to memory
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="tender.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
