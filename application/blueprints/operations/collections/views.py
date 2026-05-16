from datetime import date, datetime
from io import BytesIO
from flask import Blueprint, render_template, request, redirect, url_for, flash, abort, send_file
from flask_login import current_user
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from application.blueprints.user import login_required, roles_accepted
from application.extensions import db

from . import app_name, app_label
from .models import Collection, CollectionDetail
from ..daily_sales.models import TransactionTender
from ..bank_account.models import BankAccount
from ...register.tender.models import Tender

bp = Blueprint(app_name, __name__, template_folder="pages", url_prefix=f"/{app_name}")
ROLES_ACCEPTED = app_label


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _receivable_tenders():
    return Tender.query.filter_by(is_receivable=True).order_by(Tender.sort_order.desc()).all()


def _outstanding_lines(tender_id=None):
    """All TransactionTender lines for receivable tenders, with outstanding balance."""
    receivable_ids = {t.id for t in _receivable_tenders()}
    query = TransactionTender.query.filter(
        TransactionTender.tender_id.in_(receivable_ids)
    ).join(TransactionTender.transaction).filter(
        db.text("\"transaction\".submitted IS NOT NULL"),
        db.text("\"transaction\".cancelled IS NULL"),
    )
    if tender_id:
        query = query.filter(TransactionTender.tender_id == tender_id)
    return query.all()


def _collected(tt_id):
    """Sum already applied to a TransactionTender line."""
    result = db.session.query(
        db.func.coalesce(db.func.sum(CollectionDetail.amount_applied), 0)
    ).filter_by(transaction_tender_id=tt_id).scalar()
    return float(result)


# ---------------------------------------------------------------------------
# Receivables dashboard
# ---------------------------------------------------------------------------

@bp.route("/")
@login_required
@roles_accepted([ROLES_ACCEPTED])
def home():
    tender_id = request.args.get("tender_id", type=int)
    tenders = _receivable_tenders()

    lines = _outstanding_lines(tender_id)
    rows = []
    for tt in lines:
        collected = _collected(tt.id)
        outstanding = tt.amount - collected
        rows.append({
            "tt": tt,
            "collected": collected,
            "outstanding": outstanding,
            "is_settled": outstanding <= 0,
        })

    show_settled = request.args.get("show_settled") == "1"
    if not show_settled:
        rows = [r for r in rows if not r["is_settled"]]

    total_outstanding = sum(r["outstanding"] for r in rows if not r["is_settled"])

    context = {
        "app_label": app_label,
        "tenders": tenders,
        "selected_tender_id": tender_id,
        "rows": rows,
        "show_settled": show_settled,
        "total_outstanding": total_outstanding,
    }
    return render_template("collections/home.html", **context)


# ---------------------------------------------------------------------------
# New collection
# ---------------------------------------------------------------------------

@bp.route("/new", methods=["GET", "POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def new_collection():
    tenders = _receivable_tenders()
    bank_accounts = BankAccount.query.filter_by(active=True).order_by(BankAccount.bank_name).all()

    if request.method == "POST":
        f = request.form

        collection_date = f.get("collection_date", str(date.today()))
        tender_id       = f.get("tender_id", type=int)
        bank_account_id = f.get("bank_account_id", type=int) or None
        reference       = (f.get("reference") or "").strip()
        notes           = (f.get("notes") or "").strip()

        line_ids    = f.getlist("line_id")
        line_amounts = f.getlist("line_amount")

        if not tender_id:
            flash("Please select a tender.", "danger")
        elif not line_ids:
            flash("Please select at least one transaction line.", "danger")
        else:
            col = Collection(
                collection_date=collection_date,
                tender_id=tender_id,
                bank_account_id=bank_account_id,
                reference=reference,
                notes=notes,
                recorded_by=current_user.id,
                created_at=str(date.today()),
            )
            db.session.add(col)
            db.session.flush()

            for lid, lamt in zip(line_ids, line_amounts):
                try:
                    amt = float(lamt)
                except (ValueError, TypeError):
                    continue
                if amt <= 0:
                    continue
                detail = CollectionDetail(
                    collection_id=col.id,
                    transaction_tender_id=int(lid),
                    amount_applied=amt,
                )
                db.session.add(detail)

            db.session.commit()
            flash("Collection recorded successfully.", "success")
            return redirect(url_for(f"{app_name}.view_collection", collection_id=col.id))

    tender_id_pre = request.args.get("tender_id", type=int)
    lines = _outstanding_lines(tender_id_pre) if tender_id_pre else []
    outstanding_rows = []
    for tt in lines:
        collected = _collected(tt.id)
        outstanding = tt.amount - collected
        if outstanding > 0:
            outstanding_rows.append({"tt": tt, "outstanding": outstanding})

    context = {
        "app_label": app_label,
        "tenders": tenders,
        "bank_accounts": bank_accounts,
        "outstanding_rows": outstanding_rows,
        "selected_tender_id": tender_id_pre,
        "today": str(date.today()),
    }
    return render_template("collections/new_collection.html", **context)


# ---------------------------------------------------------------------------
# AJAX — load outstanding lines for a tender
# ---------------------------------------------------------------------------

@bp.route("/lines")
@login_required
def get_lines():
    from flask import jsonify
    tender_id = request.args.get("tender_id", type=int)
    if not tender_id:
        return jsonify([])
    lines = _outstanding_lines(tender_id)
    result = []
    for tt in lines:
        collected = _collected(tt.id)
        outstanding = round(tt.amount - collected, 2)
        if outstanding <= 0:
            continue
        t = tt.transaction
        result.append({
            "id": tt.id,
            "record_date": t.record_date or "",
            "record_number": t.record_number or "",
            "customer": t.customer.customer_name if t.customer else "",
            "amount": tt.amount,
            "collected": round(collected, 2),
            "outstanding": outstanding,
        })
    return jsonify(result)


# ---------------------------------------------------------------------------
# View collection
# ---------------------------------------------------------------------------

@bp.route("/<int:collection_id>")
@login_required
@roles_accepted([ROLES_ACCEPTED])
def view_collection(collection_id):
    col = Collection.query.get_or_404(collection_id)
    context = {
        "app_label": app_label,
        "col": col,
    }
    return render_template("collections/view_collection.html", **context)


# ---------------------------------------------------------------------------
# Collection history
# ---------------------------------------------------------------------------

@bp.route("/history")
@login_required
@roles_accepted([ROLES_ACCEPTED])
def history():
    cols = Collection.query.order_by(Collection.id.desc()).all()
    context = {
        "app_label": app_label,
        "cols": cols,
    }
    return render_template("collections/history.html", **context)


# ---------------------------------------------------------------------------
# Delete collection
# ---------------------------------------------------------------------------

@bp.route("/<int:collection_id>/delete", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def delete_collection(collection_id):
    col = Collection.query.get_or_404(collection_id)
    db.session.delete(col)
    db.session.commit()
    flash("Collection deleted.", "warning")
    return redirect(url_for(f"{app_name}.history"))


# ---------------------------------------------------------------------------
# Download Excel template
# ---------------------------------------------------------------------------

@bp.route("/template/download")
@login_required
@roles_accepted([ROLES_ACCEPTED])
def download_template():
    tender_id = request.args.get("tender_id", type=int)

    lines = _outstanding_lines(tender_id)
    rows = []
    for tt in lines:
        collected = _collected(tt.id)
        outstanding = round(tt.amount - collected, 2)
        if outstanding > 0:
            rows.append((tt, outstanding))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Collections"

    # --- styles ---
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="1a4a6b")
    locked_fill  = PatternFill("solid", fgColor="F2F2F2")
    input_fill   = PatternFill("solid", fgColor="FFFDE7")
    center       = Alignment(horizontal="center", vertical="center")
    right        = Alignment(horizontal="right")

    # --- instruction row ---
    ws.merge_cells("A1:J1")
    ws["A1"] = (
        "INSTRUCTIONS: Fill in Collection Date, Reference, and Amount Applied "
        "for each line you are settling. Leave blank to skip. Do not edit columns A–G."
    )
    ws["A1"].font = Font(italic=True, color="555555")

    # --- column headers (row 2) ---
    headers = [
        ("TT_ID",           7),
        ("Date",           12),
        ("OR #",           10),
        ("Patient",        28),
        ("Tender",         18),
        ("Amount",         12),
        ("Outstanding",    13),
        ("Collection Date",14),
        ("Reference",      18),
        ("Amount Applied", 14),
    ]
    for col_idx, (label, width) in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font     = header_font
        cell.fill     = header_fill
        cell.alignment = center
        ws.column_dimensions[cell.column_letter].width = width

    # --- data rows ---
    for row_idx, (tt, outstanding) in enumerate(rows, start=3):
        t = tt.transaction
        record_date = t.record_date or ""
        record_number = t.record_number or ""
        patient = t.customer.customer_name if t.customer else ""
        tender_name = tt.tender.tender_name if tt.tender else ""

        values = [tt.id, record_date, record_number, patient, tender_name,
                  tt.amount, outstanding, None, None, outstanding]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx <= 7:
                cell.fill      = locked_fill
                cell.alignment = right if col_idx in (6, 7) else cell.alignment
            else:
                cell.fill = input_fill
            if col_idx in (6, 7, 10):
                cell.number_format = "#,##0.00"

    # Freeze header rows
    ws.freeze_panes = "A3"

    # Lock hint: grey out ID column
    ws.column_dimensions["A"].hidden = False

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"collections_template_{date.today().isoformat()}.xlsx"
    return send_file(stream, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------------------------------------------------------------------
# Upload Excel template
# ---------------------------------------------------------------------------

@bp.route("/upload", methods=["GET", "POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def upload_collections():
    bank_accounts = BankAccount.query.filter_by(active=True).order_by(BankAccount.bank_name).all()

    if request.method == "GET":
        return render_template("collections/upload.html",
                               app_label=app_label, bank_accounts=bank_accounts)

    # ---- POST: parse file ----
    uploaded = request.files.get("xlsx_file")
    if not uploaded or not uploaded.filename.endswith(".xlsx"):
        flash("Please upload a valid .xlsx file.", "danger")
        return render_template("collections/upload.html",
                               app_label=app_label, bank_accounts=bank_accounts)

    bank_account_id = request.form.get("bank_account_id", type=int) or None
    action          = request.form.get("action", "preview")

    try:
        wb = openpyxl.load_workbook(uploaded, data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f"Could not read file: {e}", "danger")
        return render_template("collections/upload.html",
                               app_label=app_label, bank_accounts=bank_accounts)

    # Read data rows (skip rows 1 and 2 which are instruction + header)
    parsed = []
    errors = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        tt_id, rec_date, rec_num, patient, tender_name, amount, outstanding, \
            col_date, reference, amt_applied = (list(row) + [None]*10)[:10]

        if tt_id is None:
            continue
        if col_date is None or amt_applied is None:
            continue  # not filled in — skip

        # Normalise collection date
        if isinstance(col_date, datetime):
            col_date = col_date.date().isoformat()
        elif isinstance(col_date, date):
            col_date = col_date.isoformat()
        else:
            col_date = str(col_date).strip()

        try:
            amt_applied = float(amt_applied)
        except (ValueError, TypeError):
            errors.append(f"Row with TT_ID={tt_id}: invalid Amount Applied '{amt_applied}'.")
            continue

        if amt_applied <= 0:
            continue

        reference = str(reference).strip() if reference else ""

        parsed.append({
            "tt_id":      int(tt_id),
            "col_date":   col_date,
            "reference":  reference,
            "amt_applied": amt_applied,
            "patient":    patient or "",
            "rec_num":    rec_num or "",
        })

    if not parsed:
        flash("No collection rows found in the file. Make sure Collection Date and Amount Applied are filled in.", "warning")
        return render_template("collections/upload.html",
                               app_label=app_label, bank_accounts=bank_accounts)

    # Validate TransactionTender IDs
    tt_ids = [r["tt_id"] for r in parsed]
    tt_map = {tt.id: tt for tt in TransactionTender.query.filter(TransactionTender.id.in_(tt_ids)).all()}
    valid_rows = []
    for r in parsed:
        if r["tt_id"] not in tt_map:
            errors.append(f"TT_ID {r['tt_id']} not found — row skipped.")
        else:
            valid_rows.append(r)

    # Group by (col_date, reference) → collection batches
    from collections import OrderedDict
    batches = OrderedDict()
    for r in valid_rows:
        key = (r["col_date"], r["reference"])
        batches.setdefault(key, []).append(r)

    # Build preview structure
    preview = []
    for (col_date, reference), lines in batches.items():
        tender_ids = {tt_map[r["tt_id"]].tender_id for r in lines}
        tender_names = list({tt_map[r["tt_id"]].tender.tender_name for r in lines
                             if tt_map[r["tt_id"]].tender})
        preview.append({
            "col_date":    col_date,
            "reference":   reference,
            "tender_names": tender_names,
            "lines":       lines,
            "total":       sum(r["amt_applied"] for r in lines),
        })

    return render_template("collections/upload_preview.html",
                           app_label=app_label,
                           preview=preview,
                           errors=errors,
                           bank_accounts=bank_accounts,
                           bank_account_id=bank_account_id,
                           tt_map=tt_map)


# ---------------------------------------------------------------------------
# Upload DSR Excel (existing workbook, reads COLLECTION columns)
# ---------------------------------------------------------------------------

def _norm_name(name):
    """Normalise patient name for fuzzy matching."""
    if not name:
        return ""
    return " ".join(str(name).upper().split())


def _to_date_str(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    # try parsing common formats
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return s or None


def _find_collection_cols(ws):
    """
    Scan row 2 for the 'COLLECTION' label, then read row 3 headers after
    that position to locate Date / Reference / Amount columns (0-indexed).
    Returns (col_date, col_ref, col_amt) or None if not found.
    """
    max_col = ws.max_column
    row2 = [ws.cell(row=2, column=c).value for c in range(1, max_col + 1)]
    marker_col = next(
        (i for i, v in enumerate(row2) if v and "COLLECTION" in str(v).upper()),
        None
    )
    if marker_col is None:
        return None

    row3 = [ws.cell(row=3, column=c).value for c in range(1, max_col + 1)]
    col_date = col_ref = col_amt = None
    for i in range(marker_col, len(row3)):
        h = str(row3[i] or "").strip().lower()
        if h == "date" and col_date is None:
            col_date = i
        elif h == "reference" and col_ref is None:
            col_ref = i
        elif "amount" in h and col_amt is None:
            col_amt = i
    if None in (col_date, col_amt):
        return None
    return col_date, col_ref, col_amt


def _find_patient_col(row3):
    for i, v in enumerate(row3):
        if v and "patient" in str(v).lower():
            return i
    # APE uses 'Description' (company name) — fall back to col 2
    for i, v in enumerate(row3):
        if v and "description" in str(v).lower():
            return i
    return 2  # default fallback


def _find_receivable_cols(row3, receivable_tenders):
    """Map column indices to Tender objects based on header text."""
    result = {}
    for i, v in enumerate(row3):
        if not v:
            continue
        header = str(v).strip().lower()
        for t in receivable_tenders:
            if t.tender_name.lower() == header or t.tender_name.lower() in header:
                result[i] = t
                break
    return result


def _parse_dsr_sheets(wb):
    """
    Parse all sheets that contain a COLLECTION section.
    Returns list of dicts: {tx_date, patient, col_date, col_ref, col_amt,
                             tender_hint (Tender or None), sheet_name, row_num}
    """
    receivable_tenders = Tender.query.filter_by(is_receivable=True).all()
    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cols = _find_collection_cols(ws)
        if not cols:
            continue
        col_date_i, col_ref_i, col_amt_i = cols

        row3 = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
        patient_col = _find_patient_col(row3)
        receivable_cols = _find_receivable_cols(row3, receivable_tenders)

        for row_num, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            row = list(row)

            # Pad row if shorter than expected
            while len(row) <= max(col_date_i, col_amt_i):
                row.append(None)

            col_date_val = row[col_date_i]
            col_amt_val  = row[col_amt_i] if col_amt_i < len(row) else None

            if not col_date_val or not col_amt_val:
                continue

            tx_date_val = row[0]
            patient_val = row[patient_col] if patient_col < len(row) else None

            tx_date_str  = _to_date_str(tx_date_val)
            col_date_str = _to_date_str(col_date_val)
            if not tx_date_str or not col_date_str:
                continue

            try:
                amt = float(col_amt_val)
            except (ValueError, TypeError):
                continue
            if amt <= 0:
                continue

            col_ref_val = row[col_ref_i] if col_ref_i is not None and col_ref_i < len(row) else None

            # Determine which receivable tender the row belongs to
            tender_hint = None
            for col_idx, tender in receivable_cols.items():
                if col_idx < len(row) and row[col_idx]:
                    tender_hint = tender
                    break

            results.append({
                "tx_date":    tx_date_str,
                "patient":    _norm_name(patient_val),
                "col_date":   col_date_str,
                "col_ref":    str(col_ref_val).strip() if col_ref_val else "",
                "col_amt":    amt,
                "tender_hint": tender_hint,
                "sheet":      sheet_name,
                "row_num":    row_num,
            })

    return results


def _match_to_tt(tx_date, patient_norm, tender_hint):
    """
    Find a TransactionTender by transaction date + normalised patient name.
    Returns (TransactionTender, confidence) or (None, reason_string).
    """
    from ..daily_sales.models import Transaction
    txns = Transaction.query.filter(
        Transaction.record_date == tx_date,
        Transaction.submitted.isnot(None),
        Transaction.cancelled.is_(None),
    ).all()

    if not txns:
        return None, "no submitted transaction on this date"

    # Match by normalised customer name
    matched_txn = None
    for txn in txns:
        if txn.customer and _norm_name(txn.customer.customer_name) == patient_norm:
            matched_txn = txn
            break

    if not matched_txn:
        return None, f"no transaction on {tx_date} for '{patient_norm}'"

    # Find receivable TransactionTender(s)
    receivable_tts = [tt for tt in matched_txn.transaction_tenders
                      if tt.tender and tt.tender.is_receivable]

    if not receivable_tts:
        return None, f"transaction found but has no receivable tender"

    if tender_hint:
        for tt in receivable_tts:
            if tt.tender.tender_name.lower() == tender_hint.tender_name.lower():
                return tt, "ok"
        # Fallback: if only one receivable tender, use it
    if len(receivable_tts) == 1:
        return receivable_tts[0], "ok"

    return None, f"multiple receivable tenders — cannot auto-select"


@bp.route("/upload-dsr", methods=["GET", "POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def upload_dsr():
    bank_accounts = BankAccount.query.filter_by(active=True).order_by(BankAccount.bank_name).all()

    if request.method == "GET":
        return render_template("collections/upload_dsr.html",
                               app_label=app_label, bank_accounts=bank_accounts)

    uploaded = request.files.get("xlsx_file")
    if not uploaded or not uploaded.filename.endswith(".xlsx"):
        flash("Please upload a valid .xlsx file.", "danger")
        return render_template("collections/upload_dsr.html",
                               app_label=app_label, bank_accounts=bank_accounts)

    bank_account_id = request.form.get("bank_account_id", type=int) or None
    action          = request.form.get("action", "preview")

    try:
        wb = openpyxl.load_workbook(uploaded, data_only=True)
    except Exception as e:
        flash(f"Could not read file: {e}", "danger")
        return render_template("collections/upload_dsr.html",
                               app_label=app_label, bank_accounts=bank_accounts)

    parsed = _parse_dsr_sheets(wb)

    if not parsed:
        flash("No collection rows found in this file. Make sure the COLLECTION columns are filled in.", "warning")
        return render_template("collections/upload_dsr.html",
                               app_label=app_label, bank_accounts=bank_accounts)

    # Match each row to a TransactionTender
    matched   = []
    unmatched = []
    for r in parsed:
        tt, reason = _match_to_tt(r["tx_date"], r["patient"], r["tender_hint"])
        if tt:
            already = _collected(tt.id)
            outstanding = round(tt.amount - already, 2)
            matched.append({**r, "tt": tt, "outstanding": outstanding})
        else:
            unmatched.append({**r, "reason": reason})

    # Group matched rows by (col_date, col_ref) into batches
    from collections import OrderedDict
    batches = OrderedDict()
    for r in matched:
        key = (r["col_date"], r["col_ref"])
        batches.setdefault(key, []).append(r)

    preview = []
    for (col_date, col_ref), lines in batches.items():
        preview.append({
            "col_date":    col_date,
            "reference":   col_ref,
            "tender_names": list({r["tt"].tender.tender_name for r in lines if r["tt"].tender}),
            "lines":       lines,
            "total":       sum(r["col_amt"] for r in lines),
        })

    return render_template("collections/upload_dsr_preview.html",
                           app_label=app_label,
                           preview=preview,
                           unmatched=unmatched,
                           bank_accounts=bank_accounts,
                           bank_account_id=bank_account_id)


@bp.route("/upload-dsr/confirm", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def confirm_dsr_upload():
    bank_account_id = request.form.get("bank_account_id", type=int) or None
    col_dates    = request.form.getlist("col_date")
    references   = request.form.getlist("reference")
    tt_ids       = request.form.getlist("tt_id")
    amts_applied = request.form.getlist("amt_applied")

    if not tt_ids:
        flash("No data to save.", "warning")
        return redirect(url_for(f"{app_name}.upload_dsr"))

    tt_id_ints = [int(x) for x in tt_ids]
    tt_map = {tt.id: tt for tt in TransactionTender.query.filter(TransactionTender.id.in_(tt_id_ints)).all()}

    from collections import OrderedDict
    batches = OrderedDict()
    for col_date, reference, tt_id, amt in zip(col_dates, references, tt_id_ints, amts_applied):
        key = (col_date, reference)
        try:
            amt = float(amt)
        except (ValueError, TypeError):
            continue
        batches.setdefault(key, []).append({"tt_id": tt_id, "amt_applied": amt})

    created = 0
    for (col_date, reference), lines in batches.items():
        first_tt = tt_map.get(lines[0]["tt_id"])
        if not first_tt:
            continue
        col = Collection(
            collection_date=col_date,
            tender_id=first_tt.tender_id,
            bank_account_id=bank_account_id,
            reference=reference,
            recorded_by=current_user.id,
            created_at=str(date.today()),
        )
        db.session.add(col)
        db.session.flush()
        for r in lines:
            db.session.add(CollectionDetail(
                collection_id=col.id,
                transaction_tender_id=r["tt_id"],
                amount_applied=r["amt_applied"],
            ))
        created += 1

    db.session.commit()
    flash(f"{created} collection batch(es) recorded from DSR file.", "success")
    return redirect(url_for(f"{app_name}.history"))


@bp.route("/upload/confirm", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def confirm_upload():
    """Save collection batches submitted from the preview page hidden fields."""
    bank_account_id = request.form.get("bank_account_id", type=int) or None

    col_dates    = request.form.getlist("col_date")
    references   = request.form.getlist("reference")
    tt_ids       = request.form.getlist("tt_id")
    amts_applied = request.form.getlist("amt_applied")

    if not tt_ids:
        flash("No data to save.", "warning")
        return redirect(url_for(f"{app_name}.upload_collections"))

    # Re-fetch TransactionTender records
    tt_id_ints = [int(x) for x in tt_ids]
    tt_map = {tt.id: tt for tt in TransactionTender.query.filter(TransactionTender.id.in_(tt_id_ints)).all()}

    # Rebuild batches from parallel lists
    from collections import OrderedDict
    batches = OrderedDict()
    for col_date, reference, tt_id, amt in zip(col_dates, references, tt_id_ints, amts_applied):
        key = (col_date, reference)
        try:
            amt = float(amt)
        except (ValueError, TypeError):
            continue
        batches.setdefault(key, []).append({"tt_id": tt_id, "amt_applied": amt})

    created = 0
    for (col_date, reference), lines in batches.items():
        first_tt = tt_map.get(lines[0]["tt_id"])
        if not first_tt:
            continue
        col = Collection(
            collection_date=col_date,
            tender_id=first_tt.tender_id,
            bank_account_id=bank_account_id,
            reference=reference,
            recorded_by=current_user.id,
            created_at=str(date.today()),
        )
        db.session.add(col)
        db.session.flush()
        for r in lines:
            db.session.add(CollectionDetail(
                collection_id=col.id,
                transaction_tender_id=r["tt_id"],
                amount_applied=r["amt_applied"],
            ))
        created += 1

    db.session.commit()
    flash(f"{created} collection batch(es) recorded successfully.", "success")
    return redirect(url_for(f"{app_name}.history"))
