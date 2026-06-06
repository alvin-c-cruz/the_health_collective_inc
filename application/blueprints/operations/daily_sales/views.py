from datetime import date, datetime

from flask import (
    Blueprint,
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    url_for,
)
from sqlalchemy.orm import joinedload

# Centralized audit logging
from application.blueprints.audit.utils import (
    log_create as audit_log_create,
)
from application.blueprints.audit.utils import (
    log_status_change as audit_log_status_change,
)
from application.blueprints.audit.utils import (
    log_update as audit_log_update,
)
from application.blueprints.audit.utils import (
    model_to_dict,
)
from application.extensions import db, ph_today

from ...register.customer import Customer
from ...register.product_type import ProductType
from ...register.sex.models import Sex
from ...register.tender import Tender
from ...user import current_user, login_required, roles_accepted
from ..bank_account.models import BankAccount
from . import app_label, app_name
from .audit_logger import get_audit_history, log_status_change
from .forms import Form
from .models import (
    Deposit,
    DepositItem,
    FundCategory,
    FundDisbursed,
    FundReceived,
    Payee,
    PettyCashVoucher,
    ReimbursementReceived,
    ReimbursementReport,
    Transaction,
    TransactionDetail,
    TransactionTender,
    TransactionType,
)

bp = Blueprint(app_name, __name__, template_folder="pages", url_prefix=f"/{app_name}")
ROLES_ACCEPTED = app_label


def _can_transact():
    """True for SuperAdmin, Admin, and Staff — the levels that may create/edit transactions."""
    return bool(current_user.superuser or current_user.admin or current_user.staff)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _generate_record_number():
    from .models import Transaction

    last = (
        Transaction.query.filter(Transaction.record_number.isnot(None))
        .order_by(Transaction.record_number.desc())
        .first()
    )
    if last and last.record_number:
        try:
            seq = int(last.record_number) + 1
        except ValueError:
            seq = 1
    else:
        seq = 1
    return f"{seq:05d}"


# ---------------------------------------------------------------------------
# Home
# ---------------------------------------------------------------------------


@bp.route("/", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def home():
    from datetime import timedelta

    # Import at function level to avoid scoping issues with DepositItem used later
    from .models import DepositItem as DI

    today = ph_today()
    date_str = request.args.get("date", str(today))
    try:
        selected_date = date.fromisoformat(date_str)
    except ValueError:
        selected_date = today

    # Get transactions for the selected date (including drafts for this date)
    all_transactions = (
        Transaction.query.options(
            joinedload(Transaction.deposit_items).joinedload(DI.deposit)
        )
        .filter(Transaction.record_date == str(selected_date))
        .order_by(Transaction.id.desc())
        .all()
    )

    # Get deposits for the selected date (including drafts for this date)
    all_deposits = (
        Deposit.query.filter(Deposit.record_date == str(selected_date))
        .order_by(Deposit.id.desc())
        .all()
    )

    # Get collections for the selected date (including drafts for this date)
    from application.blueprints.operations.collections.models import Collection

    all_collections = (
        Collection.query.filter(Collection.collection_date == str(selected_date))
        .order_by(Collection.id.desc())
        .all()
    )

    # Get funds received for the selected date (including drafts for this date)
    all_funds_received = (
        FundReceived.query.filter(FundReceived.record_date == str(selected_date))
        .order_by(FundReceived.id.desc())
        .all()
    )

    # Get funds disbursed for the selected date (including drafts for this date)
    all_funds_disbursed = (
        FundDisbursed.query.filter(FundDisbursed.record_date == str(selected_date))
        .order_by(FundDisbursed.id.desc())
        .all()
    )

    # Get petty cash vouchers for the selected date (including drafts for this date)
    all_petty_cash_vouchers = (
        PettyCashVoucher.query.filter(
            PettyCashVoucher.record_date == str(selected_date)
        )
        .order_by(PettyCashVoucher.id.desc())
        .all()
    )

    # Get reimbursements received for the selected date
    reimbursements_received = (
        ReimbursementReceived.query.filter(
            ReimbursementReceived.record_date == str(selected_date)
        )
        .order_by(ReimbursementReceived.id.desc())
        .all()
    )

    total_sales = sum(
        sum(d.amount - d.discount for d in t.transaction_details if d.billable)
        - (t.discount or 0)
        for t in all_transactions
        if t.submitted and not t.cancelled
    )

    # Calculate cash on hand: cash sales + funds received - funds disbursed - unreimbursed expenses (cumulative up to selected date)

    # Get cash tender IDs (tenders NOT marked as receivable)
    cash_tenders = Tender.query.filter(Tender.is_receivable == False).all()
    cash_tender_ids = [t.id for t in cash_tenders]

    # 1. Cash from transactions (cumulative up to selected date) - only non-receivable tenders
    cash_from_sales = sum(
        sum(
            td.amount for td in t.transaction_tenders if td.tender_id in cash_tender_ids
        )
        for t in Transaction.query.filter(
            Transaction.record_date <= str(selected_date),
            Transaction.submitted.isnot(None),
            Transaction.cancelled.is_(None),
        ).all()
    )

    # 2. Funds received (cumulative up to selected date, posted and submitted only)
    total_funds_received = sum(
        f.amount
        for f in FundReceived.query.filter(
            FundReceived.record_date <= str(selected_date),
            FundReceived.status.in_(["posted", "submitted"]),
        ).all()
    )

    # 3. Funds disbursed (cumulative up to selected date, posted and submitted only)
    total_funds_disbursed = sum(
        f.amount
        for f in FundDisbursed.query.filter(
            FundDisbursed.record_date <= str(selected_date),
            FundDisbursed.status.in_(["posted", "submitted"]),
        ).all()
    )

    # 4. Unreimbursed petty cash expenses (cumulative up to selected date, posted and submitted only, not yet reimbursed)
    # Get all petty cash vouchers
    all_petty_cash = PettyCashVoucher.query.filter(
        PettyCashVoucher.record_date <= str(selected_date),
        PettyCashVoucher.status.in_(["posted", "submitted"]),
    ).all()

    # Get all reimbursements received (posted and submitted only)
    all_reimbursements = ReimbursementReceived.query.filter(
        ReimbursementReceived.record_date <= str(selected_date),
        ReimbursementReceived.status.in_(["posted", "submitted"]),
    ).all()

    # Calculate unreimbursed amount
    total_petty_cash_expenses = sum(pcv.amount for pcv in all_petty_cash)
    total_reimbursements = sum(r.amount for r in all_reimbursements)
    unreimbursed_expenses = total_petty_cash_expenses - total_reimbursements

    # 5. Deposits (cumulative up to selected date, posted and submitted only) - cash going OUT to bank
    # Only count deposits from cash transactions (not receivables)
    deposit_items = (
        db.session.query(DepositItem)
        .join(Deposit)
        .join(Transaction, DepositItem.transaction_id == Transaction.id)
        .join(TransactionTender, TransactionTender.transaction_id == Transaction.id)
        .filter(
            Deposit.record_date <= str(selected_date),
            Deposit.status.in_(["posted", "submitted"]),
            TransactionTender.tender_id.in_(cash_tender_ids),
        )
        .all()
    )
    total_deposits = sum(item.amount for item in deposit_items)

    # Final cash on hand calculation
    # Cash sales + Funds received - Funds disbursed - Unreimbursed expenses - Deposits
    # NOTE: Collections are NOT included because they go directly into bank accounts, not cash on hand
    cash_on_hand = (
        cash_from_sales
        + total_funds_received
        - total_funds_disbursed
        - unreimbursed_expenses
        - total_deposits
    )

    class Summary:
        pass

    summary = Summary()
    summary.total_sales = total_sales
    summary.cash_on_hand = cash_on_hand
    summary.transaction_count = len(
        [t for t in all_transactions if t.submitted and not t.cancelled]
    )

    # Calculate demographics by product type (sum sales amounts by service type)
    summary.by_type = {}
    for t in all_transactions:
        if t.submitted and not t.cancelled:
            for detail in t.transaction_details:
                if detail.billable and detail.product and detail.product.product_type:
                    service_type = detail.product.product_type.product_type_name
                    line_total = detail.amount - detail.discount
                    summary.by_type[service_type] = (
                        summary.by_type.get(service_type, 0) + line_total
                    )

    # Calculate demographics by tender (sum tender amounts)
    summary.by_tender = {}
    for t in all_transactions:
        if t.submitted and not t.cancelled:
            for tt in t.transaction_tenders:
                if tt.tender:
                    tender_name = tt.tender.tender_name
                    summary.by_tender[tender_name] = (
                        summary.by_tender.get(tender_name, 0) + tt.amount
                    )

    all_tenders = Tender.query.order_by(Tender.tender_name).all()
    txn_type_tenders = {}
    for t in all_tenders:
        if t.transaction_types:
            for tt in t.transaction_types.split(","):
                tt = tt.strip()
                if tt:
                    txn_type_tenders.setdefault(tt, []).append(t)

    transaction_types = (
        TransactionType.query.filter_by(active=True)
        .order_by(TransactionType.sort_order)
        .all()
    )

    # Build a lookup dict: type_code -> TransactionType for template use
    type_lookup = {tt.type_code: tt for tt in transaction_types}

    # Get count of pending change requests for admins
    from .change_request_models import ChangeRequest

    pending_change_requests_count = ChangeRequest.query.filter_by(
        status="pending"
    ).count()

    context = {
        "app_label": app_label,
        "today": today,
        "selected_date": selected_date,
        "is_today": selected_date == today,
        "prev_date": selected_date - timedelta(days=1),
        "next_date": selected_date + timedelta(days=1),
        "summary": summary,
        "transactions": all_transactions,  # Include both date-specific and all drafts
        "deposits": all_deposits,  # Include both date-specific and all draft deposits
        "collections": all_collections,  # Include both date-specific and all draft collections
        "funds_received": all_funds_received,  # Include both date-specific and all drafts
        "funds_disbursed": all_funds_disbursed,  # Include both date-specific and all drafts
        "petty_cash_vouchers": all_petty_cash_vouchers,  # Include both date-specific and all drafts
        "reimbursements_received": reimbursements_received,  # Only date-specific
        "txn_type_tenders": txn_type_tenders,
        "transaction_types": transaction_types,
        "type_lookup": type_lookup,
        "pending_change_requests_count": pending_change_requests_count,
    }
    return render_template("daily_sales/home.html", **context)


# ---------------------------------------------------------------------------
# Drafts - Show all draft transactions
# ---------------------------------------------------------------------------


@bp.route("/drafts", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def drafts():
    """Show all draft transactions and deposits (not submitted)"""
    from ..collections.models import Collection

    # Get all transactions where submitted is NULL (draft status)
    draft_transactions = (
        Transaction.query.filter(
            Transaction.submitted == None, Transaction.cancelled == None
        )
        .order_by(Transaction.record_date.desc(), Transaction.id.desc())
        .all()
    )

    # Get all deposits with draft status
    draft_deposits = (
        Deposit.query.filter(Deposit.status == "draft")
        .order_by(Deposit.record_date.desc(), Deposit.id.desc())
        .all()
    )

    # Get all collections with draft status
    draft_collections = (
        Collection.query.filter(Collection.status == "draft")
        .order_by(Collection.collection_date.desc(), Collection.id.desc())
        .all()
    )

    transaction_types = (
        TransactionType.query.filter_by(active=True)
        .order_by(TransactionType.sort_order)
        .all()
    )
    type_lookup = {tt.type_code: tt for tt in transaction_types}

    context = {
        "app_label": app_label,
        "transactions": draft_transactions,
        "deposits": draft_deposits,
        "collections": draft_collections,
        "type_lookup": type_lookup,
        "page_title": "Draft Records",
    }
    return render_template("daily_sales/drafts.html", **context)


# ---------------------------------------------------------------------------
# All Transactions - Show all transactions grouped by type
# ---------------------------------------------------------------------------


@bp.route("/all_transactions", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def all_transactions():
    """Show all transactions grouped by category"""
    from ..collections.models import Collection

    # Category 1: SALES - Revenue transactions
    transaction_types = (
        TransactionType.query.filter_by(active=True)
        .order_by(TransactionType.sort_order)
        .all()
    )
    sales_by_type = {}
    for tt in transaction_types:
        transactions = (
            Transaction.query.filter(Transaction.transaction_type_id == tt.id)
            .order_by(Transaction.record_date.desc(), Transaction.id.desc())
            .all()
        )
        if transactions:
            sales_by_type[tt] = transactions

    # Include transactions without a transaction type (e.g., imported CSV transactions)
    transactions_without_type = (
        Transaction.query.filter(Transaction.transaction_type_id == None)
        .order_by(Transaction.record_date.desc(), Transaction.id.desc())
        .all()
    )
    if transactions_without_type:
        # Create a pseudo transaction type for display
        from collections import namedtuple

        PseudoType = namedtuple("PseudoType", ["type_name"])
        untyped = PseudoType(type_name="Unclassified")
        sales_by_type[untyped] = transactions_without_type

    # Category 2: DEPOSITS - Bank deposit transactions
    deposits = Deposit.query.order_by(
        Deposit.record_date.desc(), Deposit.id.desc()
    ).all()

    # Category 3: COLLECTIONS - Collection transactions from customers
    collections = Collection.query.order_by(
        Collection.collection_date.desc(), Collection.id.desc()
    ).all()

    # Category 4: ACCOUNTABILITIES - Fund received and disbursed
    accountabilities_received = FundReceived.query.order_by(
        FundReceived.record_date.desc(), FundReceived.id.desc()
    ).all()
    accountabilities_disbursed = FundDisbursed.query.order_by(
        FundDisbursed.record_date.desc(), FundDisbursed.id.desc()
    ).all()

    # Category 5: PETTY CASH EXPENSES - Expense transactions from petty cash
    petty_cash_expenses = PettyCashVoucher.query.order_by(
        PettyCashVoucher.record_date.desc(), PettyCashVoucher.id.desc()
    ).all()

    # Category 6: PETTY CASH REIMBURSEMENTS - Reimbursement transactions
    petty_cash_reimbursements = ReimbursementReport.query.order_by(
        ReimbursementReport.created_date.desc(), ReimbursementReport.id.desc()
    ).all()

    context = {
        "app_label": app_label,
        "sales_by_type": sales_by_type,
        "deposits": deposits,
        "collections": collections,
        "accountabilities_received": accountabilities_received,
        "accountabilities_disbursed": accountabilities_disbursed,
        "petty_cash_expenses": petty_cash_expenses,
        "petty_cash_reimbursements": petty_cash_reimbursements,
        "page_title": "All Transactions",
    }
    return render_template("daily_sales/all_transactions.html", **context)


# ---------------------------------------------------------------------------
# New transaction  (GET: show form | POST: save)
# ---------------------------------------------------------------------------


@bp.route("/transaction/new", methods=["GET", "POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def new_transaction():
    if not _can_transact():
        abort(403)

    ape_batch_id_arg = request.args.get("ape_batch_id", type=int)
    locked_from_batch = False

    if ape_batch_id_arg:
        ape_type = TransactionType.query.filter_by(type_code="ape").first()
        transaction_type = ape_type
        locked_from_batch = True
    else:
        type_id = request.args.get("type_id", type=int)
        transaction_type = (
            TransactionType.query.get(type_id)
            if type_id
            else TransactionType.query.order_by(TransactionType.sort_order).first()
        )

    form = Form()
    form.user_prepare_id = current_user.id

    # Use date from query parameter (from daily sales home page) or default to today
    default_date = request.args.get("date", str(ph_today()))
    form.record_date = default_date

    form.record_number = _generate_record_number()
    form.transaction_type_id = transaction_type.id if transaction_type else None
    if ape_batch_id_arg:
        form.ape_batch_id = ape_batch_id_arg

    prefill_tender_id = request.args.get("prefill_tender_id", type=int)
    if prefill_tender_id:
        form.tenders[0][1].tender_id = prefill_tender_id
    elif transaction_type and transaction_type.type_code == "ape":
        # Default tender to Credit for all APE transactions
        credit_tender = Tender.query.filter(Tender.tender_name.ilike("credit")).first()
        if credit_tender:
            form.tenders[0][1].tender_id = credit_tender.id
            # If locked from a batch, also pre-fill the package amount
            if locked_from_batch and ape_batch_id_arg:
                from ..ape_batch.models import ApeBatch

                batch = ApeBatch.query.get(ape_batch_id_arg)
                if batch and batch.package_amount:
                    form.tenders[0][1].amount = batch.package_amount

    if request.method == "POST":
        form._post(request.form)
        transaction_type = (
            TransactionType.query.get(form.transaction_type_id)
            if form.transaction_type_id
            else None
        )
        if not form.record_number:
            form.record_number = _generate_record_number()

        action = request.form.get("action", "save")

        if form._validate_on_submit():
            form._save()

            if action == "submit":
                form._submit()
                flash("Transaction saved and submitted.", "success")
            else:
                flash("Transaction saved as draft.", "success")

            return redirect(
                url_for(f"{app_name}.edit_transaction", transaction_id=form.id)
            )

    product_types = ProductType.query.order_by(ProductType.product_type_name).all()
    tenders = Tender.query.order_by(Tender.tender_name).all()
    sexes = Sex.query.order_by(Sex.sex_name).all()
    customers = Customer.query.order_by(
        Customer.last_name, Customer.first_name, Customer.middle_name
    ).all()
    transaction_types = (
        TransactionType.query.filter_by(active=True)
        .order_by(TransactionType.sort_order)
        .all()
    )
    from ..ape_batch.models import ApeBatch

    ape_batches = ApeBatch.query.order_by(ApeBatch.batch_date.desc()).all()

    context = {
        "form": form,
        "transaction_type": transaction_type,
        "transaction_types": transaction_types,
        "product_types": product_types,
        "tenders": tenders,
        "sexes": sexes,
        "customers": customers,
        "ape_batches": ape_batches,
        "app_label": app_label,
        "is_new": True,
        "locked_from_batch": locked_from_batch,
    }
    return render_template("daily_sales/new_transaction.html", **context)


# ---------------------------------------------------------------------------
# Edit transaction  (GET: load | POST: update)
# ---------------------------------------------------------------------------


@bp.route("/transaction/<int:transaction_id>/edit", methods=["GET", "POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def edit_transaction(transaction_id):
    if not _can_transact():
        abort(403)
    record = Transaction.query.get_or_404(transaction_id)
    form = Form()
    form.user_prepare_id = current_user.id

    # Get return URL from query parameter or referer
    return_url = (
        request.args.get("return_url")
        or request.referrer
        or url_for("daily_sales.home")
    )

    # If return_url points to new transaction page or edit page, redirect to daily sales home instead
    # BUT: Allow returning to special pages like drafts, pending approval, etc.
    if "/transaction/new" in return_url.lower() or (
        "/edit" in return_url.lower()
        and "/drafts" not in return_url.lower()
        and "/pending" not in return_url.lower()
    ):
        return_url = url_for("daily_sales.home")

    # If return_url is daily sales home without date parameter, add the record date
    # But don't modify if it's drafts page or other special pages
    if (
        "/daily_sales/" in return_url
        and "?date=" not in return_url
        and "date=" not in return_url
        and "/drafts" not in return_url
        and "/pending" not in return_url
    ):
        return_url = url_for("daily_sales.home", date=record.record_date)

    if record.submitted or record.cancelled:
        flash("This transaction is locked and cannot be edited.", "warning")
        return redirect(
            url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
        )

    if request.method == "POST":
        form._post(request.form)
        form.id = transaction_id

        action = request.form.get("action", "save")
        cancel_action = request.form.get("cancel")  # Check if Cancel button was clicked

        # If Cancel was clicked, redirect to return URL
        if cancel_action:
            return redirect(return_url)

        if form._validate_on_submit():
            form._save()

            if action == "submit":
                form._submit()
                flash("Transaction submitted successfully.", "success")
                # After submit, go to daily sales home page with the transaction's date
                return redirect(url_for("daily_sales.home", date=record.record_date))
            flash("Transaction updated.", "success")
            # After save draft, stay on edit page
            return redirect(
                url_for(
                    f"{app_name}.edit_transaction",
                    transaction_id=form.id,
                    return_url=return_url,
                )
            )
    else:
        form._populate(record)

    transaction_type = record.transaction_type
    product_types = ProductType.query.order_by(ProductType.product_type_name).all()
    tenders = Tender.query.order_by(Tender.tender_name).all()
    sexes = Sex.query.order_by(Sex.sex_name).all()
    customers = Customer.query.order_by(
        Customer.last_name, Customer.first_name, Customer.middle_name
    ).all()
    transaction_types = (
        TransactionType.query.filter_by(active=True)
        .order_by(TransactionType.sort_order)
        .all()
    )
    from ..ape_batch.models import ApeBatch

    ape_batches = ApeBatch.query.order_by(ApeBatch.batch_date.desc()).all()

    context = {
        "form": form,
        "transaction_type": transaction_type,
        "transaction_types": transaction_types,
        "product_types": product_types,
        "tenders": tenders,
        "sexes": sexes,
        "customers": customers,
        "ape_batches": ape_batches,
        "app_label": app_label,
        "is_new": False,
        "record": record,
        "return_url": return_url,
    }
    return render_template("daily_sales/new_transaction.html", **context)


# ---------------------------------------------------------------------------
# Guide
# ---------------------------------------------------------------------------


@bp.route("/guide")
@login_required
@roles_accepted([ROLES_ACCEPTED])
def guide():
    return render_template("daily_sales/guide.html")


# ---------------------------------------------------------------------------
# View (read-only)
# ---------------------------------------------------------------------------


@bp.route("/transaction/<int:transaction_id>", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def view_transaction(transaction_id):
    record = Transaction.query.get_or_404(transaction_id)
    context = {
        "record": record,
        "app_label": app_label,
    }
    return render_template("daily_sales/view_transaction.html", **context)


# ---------------------------------------------------------------------------
# Cancel
# ---------------------------------------------------------------------------


@bp.route("/transaction/<int:transaction_id>/cancel", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def cancel_transaction(transaction_id):
    if not _can_transact():
        abort(403)
    record = Transaction.query.get_or_404(transaction_id)

    if record.cancelled:
        flash("Transaction is already cancelled.", "warning")
    elif record.submitted:
        flash(
            "Submitted transactions cannot be cancelled here. Contact an admin.",
            "danger",
        )
    else:
        form = Form()
        form.id = transaction_id
        form._populate(record)
        form._cancel()
        flash("Transaction cancelled.", "warning")

    return redirect(
        url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
    )


@bp.route("/transaction/<int:transaction_id>/uncancel", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def uncancel_transaction(transaction_id):
    """Un-cancel a transaction - only allowed before submission"""
    record = Transaction.query.get_or_404(transaction_id)

    if not record.cancelled:
        flash("Transaction is not cancelled.", "warning")
    elif record.submitted:
        flash(
            "Submitted transactions cannot be un-cancelled. Contact an admin.", "danger"
        )
    else:
        form = Form()
        form.id = transaction_id
        form._populate(record)
        success = form._uncancel()
        if success:
            flash("Transaction un-cancelled successfully.", "success")
        else:
            flash("Unable to un-cancel this transaction.", "danger")

    return redirect(
        url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
    )


# ---------------------------------------------------------------------------
# Audit History - View audit trail for a transaction
# ---------------------------------------------------------------------------


@bp.route("/transaction/<int:transaction_id>/audit_history", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def transaction_audit_history(transaction_id):
    """View complete audit history for a specific transaction"""
    from .audit_logger import get_audit_history

    record = Transaction.query.get_or_404(transaction_id)
    audit_logs = get_audit_history("transaction", transaction_id)

    context = {
        "app_label": app_label,
        "record": record,
        "audit_logs": audit_logs,
        "page_title": f"Audit History - Transaction #{record.record_number or transaction_id}",
    }
    return render_template("daily_sales/audit_history.html", **context)


# ---------------------------------------------------------------------------
# Request cancellation (staff can request cancellation of submitted transaction)
# ---------------------------------------------------------------------------


@bp.route("/transaction/<int:transaction_id>/request_cancellation", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def request_cancellation(transaction_id):
    """Staff requests cancellation of a submitted transaction"""
    from datetime import datetime

    from application.extensions import db

    record = Transaction.query.get_or_404(transaction_id)

    if not record.submitted or record.cancelled:
        flash(
            "Only submitted, active transactions can have cancellation requested.",
            "warning",
        )
        return redirect(
            url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
        )

    if record.cancellation_requested_at:
        flash("Cancellation has already been requested for this transaction.", "info")
        return redirect(
            url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
        )

    reason = request.form.get("reason", "").strip()
    if not reason:
        flash("A reason is required to request cancellation.", "danger")
        return redirect(
            url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
        )

    # Record the cancellation request
    record.cancellation_requested_by_id = current_user.id
    record.cancellation_requested_at = datetime.now()
    record.cancellation_reason = reason

    db.session.commit()

    # Audit logging
    try:
        log_status_change(
            "transaction", record, "cancellation_requested", reason=reason
        )
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f"Cancellation requested, but audit logging failed: {e!s}", "warning")
        print(f"Audit logging failed for transaction {transaction_id}: {e}")
        return redirect(
            url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
        )

    flash("Cancellation request submitted. Waiting for admin approval.", "success")
    return redirect(
        url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
    )


# ---------------------------------------------------------------------------
# Approve cancellation (admin approves cancellation request)
# ---------------------------------------------------------------------------


@bp.route("/transaction/<int:transaction_id>/approve_cancellation", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def approve_cancellation(transaction_id):
    """Admin approves cancellation request and cancels the transaction"""
    from datetime import datetime

    from flask import abort

    from application.extensions import db

    if not (current_user.admin or current_user.superuser):
        abort(403)

    record = Transaction.query.get_or_404(transaction_id)

    if not record.cancellation_requested_at:
        flash("No cancellation request found for this transaction.", "warning")
        return redirect(
            url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
        )

    if record.cancelled:
        flash("Transaction is already cancelled.", "info")
        return redirect(
            url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
        )

    # Approve cancellation and cancel the transaction
    record.cancellation_approved_by_id = current_user.id
    record.cancellation_approved_at = datetime.now()
    record.cancelled = str(datetime.now().date())

    db.session.commit()

    # Audit logging
    try:
        log_status_change(
            "transaction",
            record,
            "cancellation_approved",
            reason=record.cancellation_reason,
        )
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f"Cancellation approved, but audit logging failed: {e!s}", "warning")
        print(f"Audit logging failed for transaction {transaction_id}: {e}")
        return redirect(
            url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
        )

    flash("Cancellation approved. Transaction has been cancelled.", "success")
    return redirect(
        url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
    )


# ---------------------------------------------------------------------------
# Reject cancellation (admin rejects cancellation request)
# ---------------------------------------------------------------------------


@bp.route("/transaction/<int:transaction_id>/reject_cancellation", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def reject_cancellation(transaction_id):
    """Admin rejects cancellation request"""
    from flask import abort

    from application.extensions import db

    if not (current_user.admin or current_user.superuser):
        abort(403)

    record = Transaction.query.get_or_404(transaction_id)

    if not record.cancellation_requested_at:
        flash("No cancellation request found for this transaction.", "warning")
        return redirect(
            url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
        )

    # Clear cancellation request
    reason_text = record.cancellation_reason  # Save before clearing
    record.cancellation_requested_by_id = None
    record.cancellation_requested_at = None
    record.cancellation_reason = None

    db.session.commit()

    # Audit logging
    try:
        log_status_change(
            "transaction", record, "cancellation_rejected", reason=reason_text
        )
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f"Cancellation rejected, but audit logging failed: {e!s}", "warning")
        print(f"Audit logging failed for transaction {transaction_id}: {e}")
        return redirect(
            url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
        )

    flash("Cancellation request rejected.", "success")
    return redirect(
        url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
    )


# ---------------------------------------------------------------------------
# Bulk submit (APE batch page)
# ---------------------------------------------------------------------------


@bp.route("/transaction/bulk_submit", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def bulk_submit():
    if not _can_transact():
        abort(403)
    ids = request.form.getlist("transaction_ids")
    batch_id = request.form.get("batch_id", type=int)
    submitted_count = 0
    for tid in ids:
        record = Transaction.query.get(int(tid))
        if record and not record.submitted and not record.cancelled:
            record.submitted = str(ph_today())
            submitted_count += 1
    db.session.commit()
    if submitted_count:
        flash(f"{submitted_count} transaction(s) submitted successfully.", "success")
    else:
        flash("No eligible draft transactions were selected.", "warning")
    if batch_id:

        return redirect(url_for("ape_batch.view_batch", batch_id=batch_id))
    return redirect(url_for(f"{app_name}.home"))


# ---------------------------------------------------------------------------
# Delete (draft only)
# ---------------------------------------------------------------------------


@bp.route("/transaction/<int:transaction_id>/delete", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def delete_transaction(transaction_id):
    if not _can_transact():
        abort(403)
    record = Transaction.query.get_or_404(transaction_id)

    if record.submitted or record.cancelled:
        flash("Only draft transactions can be deleted.", "danger")
        return redirect(
            url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
        )

    TransactionDetail.query.filter_by(transaction_id=transaction_id).delete()
    TransactionTender.query.filter_by(transaction_id=transaction_id).delete()

    from .admin_models import AdminTransaction, UserTransaction

    UserTransaction.query.filter_by(transaction_id=transaction_id).delete()
    AdminTransaction.query.filter_by(transaction_id=transaction_id).delete()

    from application.extensions import db

    db.session.delete(record)
    db.session.commit()

    flash("Transaction deleted.", "success")

    # Redirect back to referrer if it's the all_transactions page, otherwise go to home
    referrer = request.referrer
    if referrer and "all_transactions" in referrer:
        return redirect(url_for(f"{app_name}.all_transactions"))
    return redirect(url_for(f"{app_name}.home"))


# ---------------------------------------------------------------------------
# Pending approval list  (admin + superuser only)
# ---------------------------------------------------------------------------


@bp.route("/pending_approval", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def pending_approval():
    from flask import abort
    from sqlalchemy import select as sa_select

    from .admin_models import AdminTransaction

    if not (current_user.admin or current_user.superuser):
        abort(403)

    # Get pending transactions
    approved_ids = sa_select(AdminTransaction.transaction_id)
    transactions = (
        Transaction.query.filter(
            Transaction.submitted.isnot(None),
            Transaction.submitted != "",
            Transaction.cancelled.is_(None),
            Transaction.id.notin_(approved_ids),
        )
        .order_by(Transaction.record_date.asc(), Transaction.id.asc())
        .all()
    )

    # Get pending deposits (submitted but not yet posted or cancelled)
    deposits = (
        Deposit.query.filter(Deposit.status == "submitted")
        .order_by(Deposit.record_date.asc(), Deposit.id.asc())
        .all()
    )

    # Get pending funds received
    funds_received = (
        FundReceived.query.filter(FundReceived.status == "submitted")
        .order_by(FundReceived.record_date.asc(), FundReceived.id.asc())
        .all()
    )

    # Get pending funds disbursed
    funds_disbursed = (
        FundDisbursed.query.filter(FundDisbursed.status == "submitted")
        .order_by(FundDisbursed.record_date.asc(), FundDisbursed.id.asc())
        .all()
    )

    # Get pending petty cash vouchers
    petty_cash_vouchers = (
        PettyCashVoucher.query.filter(PettyCashVoucher.status == "submitted")
        .order_by(PettyCashVoucher.record_date.asc(), PettyCashVoucher.id.asc())
        .all()
    )

    # Get pending collections (submitted but not yet posted or cancelled)
    from ..collections.models import Collection

    collections = (
        Collection.query.filter(Collection.status == "submitted")
        .order_by(Collection.collection_date.asc(), Collection.id.asc())
        .all()
    )

    return render_template(
        "daily_sales/pending_approval.html",
        transactions=transactions,
        deposits=deposits,
        collections=collections,
        funds_received=funds_received,
        funds_disbursed=funds_disbursed,
        petty_cash_vouchers=petty_cash_vouchers,
        app_label=app_label,
    )


# ---------------------------------------------------------------------------
# Approve transaction  (admin + superuser only)
# ---------------------------------------------------------------------------


@bp.route("/transaction/<int:transaction_id>/approve", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def approve_transaction(transaction_id):
    from flask import abort

    from application.extensions import db

    from .approval_helpers import approve_transaction_both_systems

    if not (current_user.admin or current_user.superuser):
        abort(403)

    record = Transaction.query.get_or_404(transaction_id)

    if not record.submitted or record.cancelled:
        flash("Only submitted, active transactions can be approved.", "warning")
        return redirect(
            url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
        )

    # Use helper function to update BOTH legacy and new systems
    success = approve_transaction_both_systems(record, current_user.id)

    if not success:
        flash("Transaction is already approved.", "info")
        return redirect(
            url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
        )

    db.session.commit()

    # Audit logging - ensure it's created
    try:
        log_status_change("transaction", record, "approved")
        db.session.commit()
    except Exception as e:
        # If audit logging fails, log the error and alert user
        db.session.rollback()
        flash(f"Transaction approved, but audit logging failed: {e!s}", "warning")
        print(f"Audit logging failed for transaction {transaction_id}: {e}")
        return redirect(url_for(f"{app_name}.home", date=record.record_date))

    flash("Transaction approved.", "success")
    # Redirect to daily_sales with the transaction date
    return redirect(url_for(f"{app_name}.home", date=record.record_date))


# ---------------------------------------------------------------------------
# Unapprove transaction (for testing - superuser only)
# ---------------------------------------------------------------------------


@bp.route("/transaction/<int:transaction_id>/unapprove", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def unapprove_transaction(transaction_id):
    """Superuser unapproves an approved transaction (for testing purposes)"""
    from flask import abort

    from application.extensions import db

    from .approval_helpers import unapprove_transaction_both_systems

    if not current_user.superuser:
        abort(403)

    record = Transaction.query.get_or_404(transaction_id)

    if not record.submitted or record.cancelled:
        flash("Only submitted, active transactions can be unapproved.", "warning")
        return redirect(
            url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
        )

    # Use helper function to clear BOTH legacy and new systems
    success = unapprove_transaction_both_systems(record)

    if not success:
        flash("Transaction is not approved.", "warning")
        return redirect(
            url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
        )

    db.session.commit()

    flash(
        "Transaction approval removed. Transaction is now pending approval.", "success"
    )
    return redirect(
        url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
    )


# ---------------------------------------------------------------------------
# Disapprove transaction (return to draft) - (admin + superuser only)
# ---------------------------------------------------------------------------


@bp.route("/transaction/<int:transaction_id>/disapprove", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def disapprove_transaction(transaction_id):
    """
    Admin/SuperUser disapproves a submitted transaction, returning it to draft.
    This allows the staff member to make corrections and re-submit.
    Requires a reason that will be logged for audit purposes and displayed to staff.

    Args:
        transaction_id: ID of the transaction to disapprove

    Form Data:
        reason: Required text field explaining why transaction was disapproved

    Returns:
        Redirect to pending_approval list with flash message

    Raises:
        403: If user is not admin or superuser
        404: If transaction not found
    """
    from flask import abort

    from application.extensions import db

    from .admin_models import AdminTransaction, UserTransaction

    # Permission check: Admin or SuperUser only
    if not (current_user.admin or current_user.superuser):
        abort(403)

    record = Transaction.query.get_or_404(transaction_id)

    # Get reason from form (required)
    reason = request.form.get("reason", "").strip()
    if not reason:
        flash("A reason is required to return a transaction to draft.", "danger")
        return redirect(
            url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
        )

    # Validate current state: must be submitted
    if not record.submitted:
        flash("Only submitted transactions can be returned to draft.", "warning")
        return redirect(
            url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
        )

    # Validate not cancelled
    if record.cancelled:
        flash("Cannot return a cancelled transaction to draft.", "warning")
        return redirect(
            url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
        )

    # Check if already approved
    existing_approval = AdminTransaction.query.filter_by(
        transaction_id=transaction_id
    ).first()
    if existing_approval:
        flash(
            'Transaction is already approved. Use "Unlock" to reverse approval.',
            "warning",
        )
        return redirect(
            url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
        )

    # Audit logging: Store disapproval reason in description field or add as note
    disapproval_note = f"\n\n[RETURNED TO DRAFT {datetime.now().strftime('%Y-%m-%d %H:%M')} by {current_user.user_name}]\nReason: {reason}"

    if record.description:
        record.description = record.description + disapproval_note
    else:
        record.description = disapproval_note.strip()

    # Revert to draft status
    record.submitted = None  # Clear legacy submitted field
    record.status = "draft"  # Set modern status field to draft

    # Clear UserTransaction submission record (if exists)
    UserTransaction.query.filter_by(transaction_id=transaction_id).delete()

    db.session.commit()

    # Audit logging
    try:
        log_status_change("transaction", record, "returned_to_draft", reason=reason)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(
            f"Transaction returned to draft, but audit logging failed: {e!s}",
            "warning",
        )
        print(f"Audit logging failed for transaction {transaction_id}: {e}")
        return redirect(url_for("daily_sales.home", date=record.record_date))

    # Notification via flash message (staff will see when they next login/view their transactions)
    flash(
        f"Transaction #{record.record_number or transaction_id} returned to draft. "
        f"Reason: {reason}. The staff member will be notified.",
        "warning",
    )

    # Return to daily sales with the transaction's date
    return redirect(url_for("daily_sales.home", date=record.record_date))


# ---------------------------------------------------------------------------
# Unlock transaction  (superuser only)
# ---------------------------------------------------------------------------


@bp.route("/transaction/<int:transaction_id>/unlock", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def unlock_transaction(transaction_id):
    from flask import abort

    from application.extensions import db

    from .admin_models import AdminTransaction

    if not current_user.superuser:
        abort(403)

    AdminTransaction.query.filter_by(transaction_id=transaction_id).delete()
    db.session.commit()
    flash("Transaction unlocked and returned to pending approval.", "warning")
    return redirect(
        url_for(f"{app_name}.view_transaction", transaction_id=transaction_id)
    )


# ---------------------------------------------------------------------------
# Customer autocomplete
# ---------------------------------------------------------------------------


@bp.route("/customer/search", methods=["GET"])
@login_required
def customer_search():
    from flask import jsonify
    from sqlalchemy import or_

    from ...register.customer import Customer

    q = request.args.get("q", "").strip()
    # Search across last_name, first_name, middle_name, and legacy customer_name
    results = (
        Customer.query.filter(
            or_(
                Customer.last_name.ilike(f"%{q}%"),
                Customer.first_name.ilike(f"%{q}%"),
                Customer.middle_name.ilike(f"%{q}%"),
                Customer.customer_name.ilike(f"%{q}%"),
            )
        )
        .order_by(Customer.last_name, Customer.first_name, Customer.middle_name)
        .limit(20)
        .all()
    )
    # Return formatted names: Last Name, First Name Middle Name
    return jsonify([str(c) for c in results])


@bp.route("/tender/notes", methods=["GET"])
@login_required
def get_tender_notes():
    """Get unique notes for a specific tender type"""
    from flask import jsonify

    tender_id = request.args.get("tender_id", type=int)

    if not tender_id:
        return jsonify({"notes": []})

    # Get distinct non-null side_notes for this tender, ordered by most recent
    notes = (
        db.session.query(TransactionTender.side_note)
        .filter(
            TransactionTender.tender_id == tender_id,
            TransactionTender.side_note != None,
            TransactionTender.side_note != "",
        )
        .distinct()
        .order_by(TransactionTender.side_note)
        .limit(50)
        .all()
    )

    # Extract the note strings from the query result
    note_list = [note[0] for note in notes if note[0]]

    return jsonify({"notes": note_list})


@bp.route("/get-discount-descriptions")
@login_required
def get_discount_descriptions():
    """Get unique discount descriptions from past transactions"""
    from flask import jsonify

    # Get distinct non-null discount_description values, ordered alphabetically
    # Note: This will work once we add the discount_description column to the Transaction model
    try:
        descriptions = (
            db.session.query(Transaction.discount_description)
            .filter(
                Transaction.discount_description != None,
                Transaction.discount_description != "",
            )
            .distinct()
            .order_by(Transaction.discount_description)
            .limit(50)
            .all()
        )

        # Extract the description strings from the query result
        desc_list = [desc[0] for desc in descriptions if desc[0]]
    except:
        # If column doesn't exist yet, return empty list
        desc_list = []

    return jsonify({"descriptions": desc_list})


# ---------------------------------------------------------------------------
# Record deposit
# ---------------------------------------------------------------------------


def get_undeposited_cash_transactions():
    """Get all transactions with cash payments that haven't been deposited yet"""
    from sqlalchemy import desc

    # Get submitted transactions (including both submitted and approved)
    # Exclude cancelled transactions
    transactions = (
        Transaction.query.filter(
            Transaction.submitted.isnot(
                None
            ),  # Include all submitted transactions (approved or not)
            Transaction.cancelled.is_(None),
        )
        .order_by(desc(Transaction.id))
        .all()
    )  # Use ID for ordering to avoid date comparison issues

    # Filter transactions that have cash tenders (is_receivable = False)
    undeposited = []
    for transaction in transactions:
        # Calculate cash amount for this transaction (only non-receivable tenders)
        cash_amount = sum(
            td.amount
            for td in transaction.transaction_tenders
            if td.tender and td.tender.is_receivable == False
        )

        if cash_amount <= 0:
            continue

        # Check if already deposited (count all non-cancelled deposits: draft, submitted, posted, and pending_cancellation)
        # Note: pending_cancellation deposits are still "locked" until cancellation is approved
        deposited_amount = sum(
            item.amount
            for item in transaction.deposit_items
            if item.deposit.status
            not in ["cancelled"]  # Exclude only fully cancelled deposits
        )

        remaining_cash = cash_amount - deposited_amount

        # Only include if remaining cash is greater than 0.01 (to avoid floating point precision issues)
        if remaining_cash > 0.01:
            undeposited.append(
                {
                    "transaction": transaction,
                    "cash_amount": cash_amount,
                    "deposited_amount": deposited_amount,
                    "remaining_cash": remaining_cash,
                }
            )

    return undeposited


@bp.route("/deposit/new", methods=["GET", "POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def record_deposit():
    if not _can_transact():
        abort(403)

    from flask_login import current_user

    from application.blueprints.operations.bank_account.models import BankAccount

    from .models import Deposit, DepositItem

    # Get the date parameter for redirect
    deposit_date = request.args.get("date") or request.form.get("deposit_date")

    if request.method == "POST":
        try:
            # Create new deposit
            deposit = Deposit()
            deposit.record_date = request.form.get("record_date")
            deposit.reference_number = request.form.get("reference_number", "")
            deposit.bank_account_id = (
                request.form.get("bank_account_id", type=int) or None
            )
            deposit.notes = request.form.get("notes", "")
            deposit.deductions = float(request.form.get("deductions", 0))
            deposit.deduction_details = request.form.get("deduction_details", "")
            deposit.status = "draft"  # Deposits start as draft
            deposit.created_by_id = current_user.id
            deposit.created_at = datetime.now()

            db.session.add(deposit)
            db.session.flush()  # Get deposit ID

            # Add deposit items from selected transactions
            total_amount = 0
            transaction_ids = request.form.getlist("transaction_ids")
            txn_list = []  # Build list for audit log

            for trans_id in transaction_ids:
                amount = float(request.form.get(f"amount_{trans_id}", 0))
                if amount > 0:
                    item = DepositItem()
                    item.deposit_id = deposit.id
                    item.transaction_id = int(trans_id)
                    item.amount = amount
                    db.session.add(item)
                    total_amount += amount

                    # Add to audit log list
                    txn = Transaction.query.get(int(trans_id))
                    if txn:
                        txn_list.append(
                            f"#{txn.record_number} - {txn.customer.customer_name if txn.customer else 'No customer'} (₱{amount:,.2f})"
                        )

            # Build transaction details for audit log
            txn_details = "; ".join(txn_list) if txn_list else "No transactions"

            # Log deposit creation in centralized audit system
            try:
                audit_log_create(
                    module="deposit",
                    record_id=deposit.id,
                    record_identifier=f"Deposit #{deposit.id} - {deposit.bank_account.bank_name if deposit.bank_account else 'No bank'}",
                    new_values=model_to_dict(
                        deposit,
                        [
                            "record_date",
                            "reference_number",
                            "bank_account_id",
                            "deductions",
                            "deduction_details",
                            "notes",
                            "status",
                        ],
                    ),
                    notes=f"Draft deposit created with {len(transaction_ids)} transactions, total ₱{total_amount:,.2f}. Transactions: {txn_details}",
                )
            except Exception as e:
                print(f"Audit logging failed: {e}")

            db.session.commit()

            flash(
                f"Deposit recorded successfully! Total amount: ₱{total_amount:,.2f}",
                "success",
            )

            # Always redirect back to daily_sales with the deposit's record_date
            return redirect(url_for(f"{app_name}.home", date=deposit.record_date))

        except Exception as e:
            db.session.rollback()
            flash(f"Error recording deposit: {e!s}", "danger")

            # Redirect back with date parameter if it exists
            if deposit_date:
                return redirect(
                    url_for(f"{app_name}.record_deposit", date=deposit_date)
                )
            return redirect(url_for(f"{app_name}.record_deposit"))

    # GET request - show form
    undeposited_transactions = get_undeposited_cash_transactions()
    bank_accounts = (
        BankAccount.query.filter_by(active=True).order_by(BankAccount.bank_name).all()
    )

    context = {
        "app_label": app_label,
        "undeposited_transactions": undeposited_transactions,
        "bank_accounts": bank_accounts,
        "deposit_date": deposit_date,  # Pass date for redirect
    }

    return render_template("daily_sales/record_deposit.html", **context)


# ---------------------------------------------------------------------------
# Daily report
# ---------------------------------------------------------------------------


@bp.route("/deposit/report", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def deposit_report():
    """Deposit report - summary of all deposits arranged by date"""
    from sqlalchemy import desc

    # Get date range from query params
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    # Build query - show all deposits including cancelled
    query = Deposit.query

    if start_date:
        query = query.filter(Deposit.record_date >= start_date)
    if end_date:
        query = query.filter(Deposit.record_date <= end_date)

    # Get all deposits ordered by date (most recent first)
    deposits = query.order_by(desc(Deposit.record_date), desc(Deposit.id)).all()

    # Calculate totals (exclude cancelled deposits from totals)
    total_deposits = len([d for d in deposits if d.status != "cancelled"])
    total_amount = sum(
        deposit.total_amount for deposit in deposits if deposit.status != "cancelled"
    )
    total_deductions = float(
        sum(
            (deposit.deductions or 0)
            for deposit in deposits
            if deposit.status != "cancelled"
        )
    )

    context = {
        "app_label": app_label,
        "deposits": deposits,
        "total_deposits": total_deposits,
        "total_amount": total_amount,
        "total_deductions": total_deductions,
        "start_date": start_date,
        "end_date": end_date,
    }

    return render_template(f"{app_name}/deposit_report.html", **context)


@bp.route("/deposit/view/<int:deposit_id>", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def view_deposit(deposit_id):
    """View deposit details"""
    deposit = Deposit.query.get_or_404(deposit_id)

    # Get return URL from query parameter or referer
    return_url = (
        request.args.get("return_url")
        or request.referrer
        or url_for("daily_sales.deposit_report")
    )

    # Get audit history for this deposit
    audit_history = get_audit_history("deposit", deposit_id)

    context = {
        "app_label": app_label,
        "deposit": deposit,
        "audit_history": audit_history,
        "return_url": return_url,
    }

    return render_template(f"{app_name}/view_deposit.html", **context)


@bp.route("/deposit/edit/<int:deposit_id>", methods=["GET", "POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def edit_deposit(deposit_id):
    """Edit draft deposit"""
    if not _can_transact():
        abort(403)

    from flask_login import current_user

    from application.blueprints.operations.bank_account.models import BankAccount

    from .models import Deposit, DepositItem

    deposit = Deposit.query.get_or_404(deposit_id)

    # Get return URL from query parameter or referer
    return_url = (
        request.args.get("return_url")
        or request.referrer
        or url_for("daily_sales.deposit_report")
    )

    # Only draft deposits can be edited
    if deposit.status != "draft":
        flash("Only draft deposits can be edited.", "warning")
        return redirect(url_for(f"{app_name}.view_deposit", deposit_id=deposit_id))

    # Check if user is the creator
    if deposit.created_by_id != current_user.id:
        flash("You can only edit your own deposits.", "danger")
        return redirect(url_for(f"{app_name}.view_deposit", deposit_id=deposit_id))

    if request.method == "POST":
        try:
            # Capture old values before update
            old_values = model_to_dict(
                deposit,
                [
                    "record_date",
                    "reference_number",
                    "bank_account_id",
                    "deductions",
                    "deduction_details",
                    "notes",
                    "status",
                ],
            )

            # Update deposit fields
            deposit.record_date = request.form.get("record_date")
            deposit.reference_number = request.form.get("reference_number", "")
            deposit.bank_account_id = (
                request.form.get("bank_account_id", type=int) or None
            )
            deposit.notes = request.form.get("notes", "")
            deposit.deductions = float(request.form.get("deductions", 0))
            deposit.deduction_details = request.form.get("deduction_details", "")
            deposit.updated_at = datetime.now()

            # Delete existing deposit items
            DepositItem.query.filter_by(deposit_id=deposit.id).delete()

            # Add new deposit items from selected transactions
            total_amount = 0
            transaction_ids = request.form.getlist("transaction_ids")

            for trans_id in transaction_ids:
                amount = float(request.form.get(f"amount_{trans_id}", 0))
                if amount > 0:
                    item = DepositItem()
                    item.deposit_id = deposit.id
                    item.transaction_id = int(trans_id)
                    item.amount = amount
                    db.session.add(item)
                    total_amount += amount

            # Capture new values after update
            new_values = model_to_dict(
                deposit,
                [
                    "record_date",
                    "reference_number",
                    "bank_account_id",
                    "deductions",
                    "deduction_details",
                    "notes",
                    "status",
                ],
            )

            # Log deposit update in centralized audit system
            try:
                audit_log_update(
                    module="deposit",
                    record_id=deposit.id,
                    record_identifier=f"Deposit #{deposit.id} - {deposit.bank_account.bank_name if deposit.bank_account else 'No bank'}",
                    old_values=old_values,
                    new_values=new_values,
                    notes=f"Deposit updated with {len(transaction_ids)} transactions, total ₱{total_amount:,.2f}",
                )
            except Exception as e:
                print(f"Audit logging failed: {e}")

            db.session.commit()

            flash(
                f"Deposit updated successfully! Total amount: ₱{total_amount:,.2f}",
                "success",
            )
            # Redirect to view_deposit page so user can review and submit
            return redirect(url_for(f"{app_name}.view_deposit", deposit_id=deposit.id))

        except Exception as e:
            db.session.rollback()
            flash(f"Error updating deposit: {e!s}", "danger")
            return redirect(
                url_for(
                    f"{app_name}.edit_deposit",
                    deposit_id=deposit_id,
                    return_url=return_url,
                )
            )

    # GET request - show form with existing data
    # Get undeposited transactions (excluding transactions in OTHER deposits, but include transactions in THIS deposit)
    undeposited_transactions = get_undeposited_cash_transactions()

    # Get transactions currently in this deposit (even if they're marked as deposited)
    current_deposit_transaction_ids = [
        item.transaction_id for item in deposit.deposit_items
    ]
    current_deposit_transactions = []

    for item in deposit.deposit_items:
        txn = item.transaction
        if txn:
            # Calculate cash amount for this transaction
            cash_amount = 0
            for tender in txn.transaction_tenders:
                if tender.tender and "cash" in tender.tender.tender_name.lower():
                    cash_amount += tender.amount

            # Calculate how much is deposited in OTHER deposits
            other_deposited = sum(
                di.amount
                for di in txn.deposit_items
                if di.deposit_id != deposit.id
                and di.deposit.status not in ["cancelled"]
            )

            remaining_cash = cash_amount - other_deposited

            if remaining_cash > 0:
                current_deposit_transactions.append(
                    {
                        "transaction": txn,
                        "cash_amount": cash_amount,
                        "deposited_amount": other_deposited,
                        "remaining_cash": remaining_cash,
                    }
                )

    # Merge: current deposit transactions + undeposited transactions (avoiding duplicates)
    all_available_transactions = current_deposit_transactions.copy()
    for undeposited in undeposited_transactions:
        if undeposited["transaction"].id not in current_deposit_transaction_ids:
            all_available_transactions.append(undeposited)

    bank_accounts = (
        BankAccount.query.filter_by(active=True).order_by(BankAccount.bank_name).all()
    )

    # Get currently selected transaction IDs
    selected_transaction_ids = current_deposit_transaction_ids

    context = {
        "app_label": app_label,
        "deposit": deposit,
        "undeposited_transactions": all_available_transactions,
        "bank_accounts": bank_accounts,
        "selected_transaction_ids": selected_transaction_ids,
        "return_url": return_url,
    }

    return render_template("daily_sales/edit_deposit.html", **context)


@bp.route("/deposit/<int:deposit_id>/audit_history", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def deposit_audit_history(deposit_id):
    """View complete audit history for a specific deposit"""
    deposit = Deposit.query.get_or_404(deposit_id)
    audit_logs = get_audit_history("deposit", deposit_id)

    context = {
        "app_label": app_label,
        "deposit": deposit,
        "audit_logs": audit_logs,
        "page_title": f"Audit History - Deposit #{deposit.id}",
    }
    return render_template("daily_sales/deposit_audit_history.html", **context)


@bp.route("/deposit/submit/<int:deposit_id>", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def submit_deposit(deposit_id):
    """Submit deposit for approval"""
    deposit = Deposit.query.get_or_404(deposit_id)

    # Get return URL from form or referrer
    return_url = (
        request.form.get("return_url")
        or request.referrer
        or url_for(f"{app_name}.deposit_report")
    )

    # Only draft deposits can be submitted
    if deposit.status != "draft":
        flash("Only draft deposits can be submitted.", "danger")
        return redirect(return_url)

    # Check if user is the creator
    if deposit.created_by_id != current_user.id:
        flash("You can only submit your own deposits.", "danger")
        return redirect(return_url)

    # Always require submission + approval workflow (even for admins)
    deposit.status = "submitted"
    deposit.submitted_by_id = current_user.id
    deposit.submitted_at = datetime.now()
    deposit.updated_at = datetime.now()

    # Log in centralized audit system
    try:
        # Build transaction list for notes
        txn_list = []
        for item in deposit.deposit_items:
            txn = item.transaction
            txn_list.append(
                f"#{txn.record_number} - {txn.customer.customer_name if txn.customer else 'No customer'} (₱{item.formatted_amount})"
            )

        txn_details = "; ".join(txn_list) if txn_list else "No transactions"

        # Build context values to show in audit log
        context_values = {
            "bank": deposit.bank_account.bank_name
            if deposit.bank_account
            else "No bank",
            "deposit_date": deposit.record_date,
            "reference_number": deposit.reference_number or "N/A",
            "total_amount": f"₱{deposit.formatted_total_amount}",
        }

        audit_log_status_change(
            module="deposit",
            record_id=deposit.id,
            record_identifier=f"Deposit #{deposit.id} - {deposit.bank_account.bank_name if deposit.bank_account else 'No bank'}",
            action="submitted",
            old_status="draft",
            new_status="submitted",
            context_values=context_values,
            notes=f"Deposit submitted for approval, total ₱{deposit.formatted_total_amount}. Transactions: {txn_details}",
        )
    except Exception as e:
        print(f"Audit logging failed: {e}")

    db.session.commit()

    # Log audit trail (old system)
    log_status_change("deposit", deposit, "submitted", user=current_user)

    flash(
        f"Deposit submitted successfully! Waiting for admin approval. Total amount: ₱{deposit.formatted_total_amount}",
        "success",
    )

    # Redirect to daily_sales with the deposit date
    return redirect(url_for(f"{app_name}.home", date=deposit.record_date))


@bp.route("/transaction/reject/<int:transaction_id>", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def reject_transaction(transaction_id):
    """Reject submitted transaction - send back to draft"""
    transaction = Transaction.query.get_or_404(transaction_id)

    # Only submitted transactions can be rejected
    if not transaction.submitted:
        flash("Only submitted transactions can be rejected.", "danger")
        return redirect(url_for("daily_sales.change_requests_list"))

    # Clear submitted status to send back to draft
    transaction.submitted = None
    transaction.updated_at = datetime.now()

    db.session.commit()

    flash(f"Transaction #{transaction.id} rejected and sent back to draft.", "warning")
    return redirect(url_for("daily_sales.change_requests_list"))


@bp.route("/deposit/approve/<int:deposit_id>", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def approve_deposit(deposit_id):
    """Approve submitted deposit (admin only)"""
    deposit = Deposit.query.get_or_404(deposit_id)

    # Only submitted deposits can be approved
    if deposit.status != "submitted":
        flash("Only submitted deposits can be approved.", "danger")
        return redirect(url_for(f"{app_name}.deposit_report"))

    deposit.status = "posted"
    deposit.approved_by_id = current_user.id
    deposit.approved_at = datetime.now()
    deposit.updated_at = datetime.now()

    # Log in centralized audit system
    try:
        # Build transaction list for notes
        txn_list = []
        for item in deposit.deposit_items:
            txn = item.transaction
            txn_list.append(
                f"#{txn.record_number} - {txn.customer.customer_name if txn.customer else 'No customer'} (₱{item.formatted_amount})"
            )

        txn_details = "; ".join(txn_list) if txn_list else "No transactions"

        # Build context values to show in audit log
        context_values = {
            "bank": deposit.bank_account.bank_name
            if deposit.bank_account
            else "No bank",
            "deposit_date": deposit.record_date,
            "reference_number": deposit.reference_number or "N/A",
            "total_amount": f"₱{deposit.formatted_total_amount}",
        }

        audit_log_status_change(
            module="deposit",
            record_id=deposit.id,
            record_identifier=f"Deposit #{deposit.id} - {deposit.bank_account.bank_name if deposit.bank_account else 'No bank'}",
            action="approved",
            old_status="submitted",
            new_status="posted",
            context_values=context_values,
            notes=f"Deposit approved and posted, total ₱{deposit.formatted_total_amount}. Transactions: {txn_details}",
        )
    except Exception as e:
        print(f"Audit logging failed: {e}")

    db.session.commit()

    # Log audit trail (old system)
    log_status_change("deposit", deposit, "approved", user=current_user)

    flash(
        f"Deposit approved successfully! Total amount: ₱{deposit.formatted_total_amount}",
        "success",
    )
    return redirect(url_for(f"{app_name}.deposit_report"))


@bp.route("/deposit/disapprove/<int:deposit_id>", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def disapprove_deposit(deposit_id):
    """Disapprove submitted deposit - send back to draft (from pending approval)"""
    from flask import abort

    if not (current_user.admin or current_user.superuser):
        abort(403)

    deposit = Deposit.query.get_or_404(deposit_id)

    # Only submitted deposits can be disapproved
    if deposit.status != "submitted":
        flash("Only submitted deposits can be disapproved.", "danger")
        return redirect(url_for(f"{app_name}.pending_approval"))

    # Send back to draft
    deposit.status = "draft"
    deposit.submitted_by_id = None
    deposit.submitted_at = None
    deposit.updated_at = datetime.now()

    db.session.commit()

    # Log audit trail
    log_status_change(
        "deposit",
        deposit,
        "disapproved",
        user=current_user,
        notes="Deposit returned to draft status",
    )

    flash("Deposit disapproved and returned to draft.", "warning")
    return redirect(url_for(f"{app_name}.pending_approval"))


@bp.route("/deposit/reject/<int:deposit_id>", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def reject_deposit(deposit_id):
    """Reject submitted deposit - send back to draft"""
    deposit = Deposit.query.get_or_404(deposit_id)

    # Only submitted deposits can be rejected
    if deposit.status != "submitted":
        flash("Only submitted deposits can be rejected.", "danger")
        return redirect(url_for("daily_sales.change_requests_list"))

    # Send back to draft
    deposit.status = "draft"
    deposit.updated_at = datetime.now()

    db.session.commit()

    flash(f"Deposit #{deposit.id} rejected and sent back to draft.", "warning")
    return redirect(url_for("daily_sales.change_requests_list"))


@bp.route("/deposit/cancel/<int:deposit_id>", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def cancel_deposit(deposit_id):
    """Cancel draft deposit (soft delete - marks as cancelled)"""
    deposit = Deposit.query.get_or_404(deposit_id)

    # Only draft deposits can be cancelled
    if deposit.status != "draft":
        flash("Only draft deposits can be cancelled.", "danger")
        return redirect(url_for(f"{app_name}.deposit_report"))

    # Check if user is the creator
    if deposit.created_by_id != current_user.id:
        flash("You can only cancel your own deposits.", "danger")
        return redirect(url_for(f"{app_name}.deposit_report"))

    # Get cancellation reason from form
    cancellation_reason = request.form.get("cancellation_reason", "").strip()

    if not cancellation_reason:
        flash("Cancellation reason is required.", "danger")
        return redirect(url_for(f"{app_name}.deposit_report"))

    # Soft delete - mark as cancelled instead of deleting
    deposit.status = "cancelled"
    deposit.cancelled_by_id = current_user.id
    deposit.cancelled_at = datetime.now()
    deposit.cancellation_reason = cancellation_reason
    deposit.updated_at = datetime.now()

    # Log in centralized audit system
    try:
        audit_log_status_change(
            module="deposit",
            record_id=deposit.id,
            record_identifier=f"Deposit #{deposit.id} - {deposit.bank_account.bank_name if deposit.bank_account else 'No bank'}",
            action="cancelled",
            old_status="draft",
            new_status="cancelled",
            reason=cancellation_reason,
            notes=f"Deposit cancelled: {cancellation_reason}",
        )
    except Exception as e:
        print(f"Audit logging failed: {e}")

    db.session.commit()

    # Log audit trail (old system)
    log_status_change(
        "deposit", deposit, "cancelled", user=current_user, reason=cancellation_reason
    )

    flash("Deposit cancelled successfully.", "info")
    # Redirect to daily_sales with the deposit date
    return redirect(url_for(f"{app_name}.home", date=deposit.record_date))


# ---------------------------------------------------------------------------
# Change Requests for Deposits
# ---------------------------------------------------------------------------


@bp.route("/deposit/request-change/<int:deposit_id>", methods=["GET", "POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def request_deposit_change(deposit_id):
    """Staff/supervisor requests change to a posted deposit"""
    from .change_request_models import ChangeRequest

    deposit = Deposit.query.get_or_404(deposit_id)

    # Only submitted or posted deposits can have change requests
    if deposit.status not in ["submitted", "posted"]:
        flash("Only submitted or posted deposits can have change requests.", "danger")
        return redirect(url_for(f"{app_name}.view_deposit", deposit_id=deposit_id))

    if request.method == "POST":
        # Capture old values
        old_values = {
            "record_date": deposit.record_date,
            "bank_account": deposit.bank_account,
            "reference_number": deposit.reference_number,
            "notes": deposit.notes,
        }

        # Capture new values from form
        new_values = {
            "record_date": request.form.get("record_date", ""),
            "bank_account": request.form.get("bank_account", ""),
            "reference_number": request.form.get("reference_number", ""),
            "notes": request.form.get("notes", ""),
        }

        reason = request.form.get("reason", "").strip()

        if not reason:
            flash("Please provide a reason for the change request.", "danger")
            return redirect(
                url_for(f"{app_name}.request_deposit_change", deposit_id=deposit_id)
            )

        # Create change request
        cr = ChangeRequest()
        cr.record_type = "deposit"
        cr.record_id = deposit_id
        cr.old_values = old_values
        cr.new_values = new_values
        cr.reason = reason
        cr.status = "pending"
        cr.requested_by_id = current_user.id
        cr.requested_at = datetime.now()

        db.session.add(cr)
        db.session.commit()

        flash(
            "Change request submitted successfully. Waiting for admin approval.",
            "success",
        )
        return redirect(url_for(f"{app_name}.view_deposit", deposit_id=deposit_id))

    # GET request - show form
    # Get undeposited transaction tenders for potential addition
    from .models import DepositItem, Transaction, TransactionTender

    # Get IDs of transactions already in ANY deposit
    deposited_transaction_ids = (
        db.session.query(DepositItem.transaction_id).distinct().all()
    )
    deposited_transaction_ids = [tid[0] for tid in deposited_transaction_ids]

    # Query for ALL undeposited tenders (not just same date)
    undeposited_tenders = (
        TransactionTender.query.join(Transaction)
        .filter(
            TransactionTender.tender_id.in_([1, 2]),  # Cash and checks typically
            ~Transaction.id.in_(deposited_transaction_ids)
            if deposited_transaction_ids
            else True,  # Not in any deposit
            Transaction.submitted != None,
            Transaction.cancelled == None,
        )
        .order_by(Transaction.record_date.desc())
        .all()
    )

    context = {
        "app_label": app_label,
        "deposit": deposit,
        "undeposited_tenders": undeposited_tenders,
    }

    return render_template(f"{app_name}/request_deposit_change.html", **context)


@bp.route("/deposit/request-cancellation/<int:deposit_id>", methods=["GET", "POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def request_deposit_cancellation(deposit_id):
    """Staff/supervisor requests cancellation of a posted deposit"""
    from .change_request_models import ChangeRequest

    deposit = Deposit.query.get_or_404(deposit_id)

    # Can only request cancellation on posted deposits
    if deposit.status != "posted":
        flash("Can only request cancellation on posted deposits.", "warning")
        return redirect(url_for(f"{app_name}.view_deposit", deposit_id=deposit_id))

    # Check if there's already a pending cancellation request
    existing_request = ChangeRequest.query.filter_by(
        record_type="deposit",
        record_id=deposit.id,
        request_action="cancellation",
        status="pending",
    ).first()

    if existing_request:
        flash(
            "There is already a pending cancellation request for this deposit.",
            "warning",
        )
        return redirect(url_for(f"{app_name}.view_deposit", deposit_id=deposit_id))

    if request.method == "POST":
        reason = request.form.get("reason", "").strip()

        if not reason:
            flash("Please provide a reason for the cancellation request.", "warning")
            context = {
                "app_label": app_label,
                "deposit": deposit,
            }
            return render_template(
                f"{app_name}/request_deposit_cancellation.html", **context
            )

        # Create cancellation request
        cr = ChangeRequest()
        cr.record_type = "deposit"
        cr.record_id = deposit_id
        cr.request_action = "cancellation"
        cr.reason = reason
        cr.status = "pending"
        cr.requested_by_id = current_user.id
        cr.requested_at = datetime.now()
        cr.old_values = {}
        cr.new_values = {}

        db.session.add(cr)
        db.session.commit()

        flash("Cancellation request submitted. An admin will review it.", "success")
        return redirect(url_for(f"{app_name}.view_deposit", deposit_id=deposit_id))

    context = {
        "app_label": app_label,
        "deposit": deposit,
    }

    return render_template(f"{app_name}/request_deposit_cancellation.html", **context)


@bp.route("/deposit/change-requests", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def deposit_change_requests():
    """Admin views pending deposit change requests"""
    from .change_request_models import ChangeRequest

    # Get all pending change requests for deposits
    change_requests = (
        ChangeRequest.query.filter(
            ChangeRequest.record_type == "deposit", ChangeRequest.status == "pending"
        )
        .order_by(ChangeRequest.requested_at.desc())
        .all()
    )

    context = {
        "app_label": app_label,
        "change_requests": change_requests,
    }

    return render_template(f"{app_name}/deposit_change_requests.html", **context)


@bp.route("/deposit/change-requests/<int:cr_id>/review", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def review_deposit_change_request(cr_id):
    """Admin approves or rejects deposit change request"""
    from .change_request_models import ChangeRequest

    cr = ChangeRequest.query.get_or_404(cr_id)

    # Only pending requests can be reviewed
    if cr.status != "pending":
        flash("This change request has already been reviewed.", "warning")
        return redirect(url_for(f"{app_name}.deposit_change_requests"))

    action = request.form.get("action")  # 'approve' or 'reject'
    review_notes = request.form.get("review_notes", "").strip()

    if action == "approve":
        # Apply the changes to the deposit
        deposit = Deposit.query.get(cr.record_id)
        if deposit:
            nv = cr.new_values
            deposit.record_date = nv.get("record_date") or deposit.record_date
            deposit.bank_account = nv.get("bank_account")
            deposit.reference_number = nv.get("reference_number")
            deposit.notes = nv.get("notes")
            deposit.updated_at = datetime.now()

            # Set deposit back to submitted status for re-approval
            deposit.status = "submitted"
            deposit.submitted_by_id = current_user.id
            deposit.submitted_at = datetime.now()

        cr.status = "approved"
        flash(
            "Change request approved. Deposit updated and set to submitted status for re-approval.",
            "success",
        )

    elif action == "reject":
        cr.status = "rejected"
        flash("Change request rejected.", "info")

    else:
        flash("Invalid action.", "danger")
        return redirect(url_for(f"{app_name}.deposit_change_requests"))

    # Record review details
    cr.reviewed_by_id = current_user.id
    cr.reviewed_at = datetime.now()
    cr.review_notes = review_notes

    db.session.commit()

    return redirect(url_for(f"{app_name}.deposit_change_requests"))


@bp.route("/daily_report", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def daily_report():
    from datetime import timedelta

    report_date_str = request.args.get("date", str(ph_today()))
    try:
        curr_date = date.fromisoformat(report_date_str)
    except ValueError:
        curr_date = ph_today()

    prev_date = curr_date - timedelta(days=1)

    def build_report(target_date):
        transactions = Transaction.query.filter(
            Transaction.record_date == str(target_date),
            Transaction.submitted.isnot(None),
            Transaction.cancelled.is_(None),
        ).all()

        report = _Report()
        report.report_date = target_date

        for t in transactions:
            tx_type = t.transaction_type.type_code if t.transaction_type else "walk_in"
            if tx_type not in report.sales:
                report.sales[tx_type] = {}
            for tender in t.transaction_tenders:
                t_name = tender.tender.tender_name if tender.tender else "Unknown"
                report.sales[tx_type][t_name] = (
                    report.sales[tx_type].get(t_name, 0) + tender.amount
                )

        return report

    prev_report = build_report(prev_date)
    curr_report = build_report(curr_date)

    # Load all active transaction types — highest sort_order first (newest = top)
    transaction_types = (
        TransactionType.query.filter_by(active=True)
        .order_by(TransactionType.sort_order.desc())
        .all()
    )

    all_tenders = Tender.query.order_by(Tender.sort_order.desc()).all()
    tender_order = {t.tender_name: t.sort_order for t in all_tenders}

    def static_for(type_code):
        return [
            t.tender_name
            for t in all_tenders
            if t.report_static
            and t.transaction_types
            and type_code in t.transaction_types.split(",")
        ]

    def ordered_keys(*dicts, static=None):
        keys = set()
        for d in dicts:
            keys.update(d.keys())
        if static:
            keys.update(static)
        return sorted(keys, key=lambda k: tender_order.get(k, 0), reverse=True)

    # Build one section entry per active transaction type
    report_sections = []
    for tt in transaction_types:
        prev_sales = prev_report.sales.get(tt.type_code, {})
        curr_sales = curr_report.sales.get(tt.type_code, {})
        tenders = ordered_keys(prev_sales, curr_sales, static=static_for(tt.type_code))

        report_sections.append(
            {
                "type_code": tt.type_code,
                "type_name": tt.type_name,
                "tenders": tenders,
                "is_dialysis": tt.type_code == "dialysis",
            }
        )

    # Cash on hand = CASH tender across diagnostic sections only (dialysis is Philhealth)
    def cash_total(report):
        return sum(
            sales.get("Cash", 0.0)
            for code, sales in report.sales.items()
            if code != "dialysis"
        )

    # Calculate fund movements for display in "Cash on Hand" section (daily amounts)
    def calculate_daily_fund_movements(target_date):
        """Calculate fund received and disbursed for a specific date only"""
        funds_received = FundReceived.query.filter(
            FundReceived.record_date == str(target_date),
            FundReceived.status.in_(["submitted", "posted"]),
        ).all()

        funds_disbursed = FundDisbursed.query.filter(
            FundDisbursed.record_date == str(target_date),
            FundDisbursed.status.in_(["submitted", "posted"]),
        ).all()

        total_received = sum(f.amount for f in funds_received)
        total_disbursed = sum(f.amount for f in funds_disbursed)

        return {
            "total_received": total_received,
            "total_disbursed": total_disbursed,
        }

    # Calculate running balances for "Accountabilities" section (cumulative up to date)
    def calculate_fund_running_balance(target_date):
        """Calculate running balance for each fund category as of target_date"""
        # Get all fund transactions up to and including target_date
        funds_received = FundReceived.query.filter(
            FundReceived.record_date <= str(target_date),
            FundReceived.status.in_(["submitted", "posted"]),
        ).all()

        funds_disbursed = FundDisbursed.query.filter(
            FundDisbursed.record_date <= str(target_date),
            FundDisbursed.status.in_(["submitted", "posted"]),
        ).all()

        # Calculate running balance by category
        balance_by_category = {}
        for fr in funds_received:
            cat_name = fr.fund_category.category_name if fr.fund_category else "Unknown"
            balance_by_category[cat_name] = (
                balance_by_category.get(cat_name, 0) + fr.amount
            )

        for fd in funds_disbursed:
            cat_name = fd.fund_category.category_name if fd.fund_category else "Unknown"
            balance_by_category[cat_name] = (
                balance_by_category.get(cat_name, 0) - fd.amount
            )

        return balance_by_category

    # Calculate undeposited cash sales running balance
    def calculate_undeposited_cash(target_date):
        """Calculate running balance of undeposited cash sales as of target_date"""
        # Get all cash sales up to and including target_date
        cash_sales = Transaction.query.filter(
            Transaction.record_date <= str(target_date),
            Transaction.submitted.isnot(None),
            Transaction.cancelled.is_(None),
        ).all()

        total_cash = 0
        for txn in cash_sales:
            for tender in txn.transaction_tenders:
                if tender.tender and "cash" in tender.tender.tender_name.lower():
                    # Only count diagnostics cash (not dialysis)
                    if (
                        txn.transaction_type
                        and txn.transaction_type.type_code != "dialysis"
                    ):
                        total_cash += tender.amount

        # Subtract all submitted/posted deposits up to and including target_date
        deposits = Deposit.query.filter(
            Deposit.record_date <= str(target_date),
            Deposit.status.in_(["submitted", "posted"]),
        ).all()

        total_deposited = sum(d.total_amount for d in deposits)

        return total_cash - total_deposited

    # Calculate cash deposited for a specific date
    def calculate_cash_deposited(target_date):
        """Calculate total cash deposited on a specific date"""
        deposits = Deposit.query.filter(
            Deposit.record_date == str(target_date),
            Deposit.status.in_(["submitted", "posted"]),
        ).all()
        return sum(d.total_amount for d in deposits)

    # Calculate beginning cash balance (ending cash from previous date)
    def calculate_cash_on_hand(target_date):
        """
        Calculate cash on hand balance as of target_date
        Formula: Fund balances + Undeposited cash sales
        """
        fund_balances = calculate_fund_running_balance(target_date)
        undeposited = calculate_undeposited_cash(target_date)
        total_funds = sum(fund_balances.values())
        return total_funds + undeposited

    prev_daily_funds = calculate_daily_fund_movements(prev_date)
    curr_daily_funds = calculate_daily_fund_movements(curr_date)

    prev_fund_balances = calculate_fund_running_balance(prev_date)
    curr_fund_balances = calculate_fund_running_balance(curr_date)

    prev_undeposited = calculate_undeposited_cash(prev_date)
    curr_undeposited = calculate_undeposited_cash(curr_date)

    # Calculate cash deposited on each date
    prev_cash_deposited = calculate_cash_deposited(prev_date)
    curr_cash_deposited = calculate_cash_deposited(curr_date)

    # Calculate beginning and ending cash
    # Beginning cash for prev_date = ending cash from day before prev_date
    from datetime import timedelta

    day_before_prev = prev_date - timedelta(days=1)
    prev_beginning_cash = calculate_cash_on_hand(day_before_prev)
    prev_ending_cash = calculate_cash_on_hand(prev_date)

    # Beginning cash for curr_date = ending cash from prev_date
    curr_beginning_cash = prev_ending_cash
    curr_ending_cash = calculate_cash_on_hand(curr_date)

    # Get all fund categories for display
    fund_categories = (
        FundCategory.query.filter_by(active=True)
        .order_by(FundCategory.sort_order)
        .all()
    )

    context = {
        "prev_report": prev_report,
        "curr_report": curr_report,
        "report_sections": report_sections,
        "prev_cash": cash_total(prev_report),
        "curr_cash": cash_total(curr_report),
        "prev_daily_funds": prev_daily_funds,
        "curr_daily_funds": curr_daily_funds,
        "prev_fund_balances": prev_fund_balances,
        "curr_fund_balances": curr_fund_balances,
        "prev_undeposited": prev_undeposited,
        "curr_undeposited": curr_undeposited,
        "prev_cash_deposited": prev_cash_deposited,
        "curr_cash_deposited": curr_cash_deposited,
        "prev_beginning_cash": prev_beginning_cash,
        "curr_beginning_cash": curr_beginning_cash,
        "prev_ending_cash": prev_ending_cash,
        "curr_ending_cash": curr_ending_cash,
        "fund_categories": fund_categories,
        "app_label": app_label,
        "report_date": curr_date,
        "prev_date": prev_date,
        "next_date": curr_date + timedelta(days=1),
    }
    return render_template("daily_sales/daily_sales_report.html", **context)


@bp.route("/daily_report_details", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def daily_report_details():
    """Page 2: Detailed transaction list, deposits, and expenses"""
    from datetime import timedelta

    report_date_str = request.args.get("date", str(ph_today()))
    try:
        report_date = date.fromisoformat(report_date_str)
    except ValueError:
        report_date = ph_today()

    # Get all submitted transactions for the date
    transactions = (
        Transaction.query.filter(
            Transaction.record_date == str(report_date),
            Transaction.submitted.isnot(None),
            Transaction.cancelled.is_(None),
        )
        .order_by(Transaction.id)
        .all()
    )

    # Group transactions by tender
    transactions_by_tender = {}
    for txn in transactions:
        for tender_record in txn.transaction_tenders:
            tender_name = (
                tender_record.tender.tender_name if tender_record.tender else "Unknown"
            )
            if tender_name not in transactions_by_tender:
                transactions_by_tender[tender_name] = []

            # Calculate transaction total (only billable items)
            total = sum(
                detail.amount - detail.discount
                for detail in txn.transaction_details
                if detail.billable
            )
            total -= txn.discount or 0

            transactions_by_tender[tender_name].append(
                {
                    "transaction": txn,
                    "tender_amount": tender_record.amount,
                    "total": total,
                    "type_name": txn.transaction_type.type_name
                    if txn.transaction_type
                    else "Unclassified",
                }
            )

    # Calculate totals by tender
    tender_totals = {}
    for tender_name, txn_list in transactions_by_tender.items():
        tender_totals[tender_name] = sum(item["tender_amount"] for item in txn_list)

    # Get deposits for the date
    deposits = (
        Deposit.query.filter(
            Deposit.record_date == str(report_date),
            Deposit.status.in_(["submitted", "posted"]),
        )
        .order_by(Deposit.id)
        .all()
    )

    total_deposited = sum(d.total_amount for d in deposits)

    # Get funds received
    funds_received = (
        FundReceived.query.filter(
            FundReceived.record_date == str(report_date),
            FundReceived.status.in_(["submitted", "posted"]),
        )
        .order_by(FundReceived.id)
        .all()
    )

    total_funds_received = sum(f.amount for f in funds_received)

    # Get expenses (fund disbursed + petty cash vouchers)
    fund_disbursed = (
        FundDisbursed.query.filter(
            FundDisbursed.record_date == str(report_date),
            FundDisbursed.status.in_(["submitted", "posted"]),
        )
        .order_by(FundDisbursed.id)
        .all()
    )

    petty_cash = (
        PettyCashVoucher.query.filter(
            PettyCashVoucher.record_date == str(report_date),
            PettyCashVoucher.status.in_(["submitted", "posted"]),
        )
        .order_by(PettyCashVoucher.id)
        .all()
    )

    total_expenses = sum(f.amount for f in fund_disbursed) + sum(
        p.amount for p in petty_cash
    )

    # Sort tenders by sort_order
    all_tenders = Tender.query.order_by(Tender.sort_order.desc()).all()
    tender_order = {t.tender_name: t.sort_order for t in all_tenders}
    sorted_tender_names = sorted(
        transactions_by_tender.keys(),
        key=lambda k: tender_order.get(k, 0),
        reverse=True,
    )

    context = {
        "app_label": app_label,
        "report_date": report_date,
        "prev_date": report_date - timedelta(days=1),
        "next_date": report_date + timedelta(days=1),
        "transactions_by_tender": transactions_by_tender,
        "sorted_tender_names": sorted_tender_names,
        "tender_totals": tender_totals,
        "deposits": deposits,
        "total_deposited": total_deposited,
        "funds_received": funds_received,
        "total_funds_received": total_funds_received,
        "fund_disbursed": fund_disbursed,
        "petty_cash": petty_cash,
        "total_expenses": total_expenses,
    }
    return render_template("daily_sales/daily_report_details.html", **context)


@bp.route("/sales_summary_report", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def sales_summary_report():
    """
    Sales Summary Report - categorized by service type (Dialysis/Diagnostics)
    and transaction type (Walk-in, Home Service, etc.) with date range picker
    """
    from ...register.product import Product

    # Get date range from query params or default to current month
    today = date.today()
    start_date_str = request.args.get(
        "start_date", str(date(today.year, today.month, 1))
    )
    end_date_str = request.args.get("end_date", str(today))

    try:
        start_date = date.fromisoformat(start_date_str)
        end_date = date.fromisoformat(end_date_str)
    except ValueError:
        start_date = date(today.year, today.month, 1)
        end_date = today

    # Ensure start_date <= end_date
    if start_date > end_date:
        start_date, end_date = end_date, start_date

    # Query transactions in date range
    transactions = Transaction.query.filter(
        Transaction.record_date >= str(start_date),
        Transaction.record_date <= str(end_date),
        Transaction.submitted.isnot(None),
        Transaction.cancelled.is_(None),
    ).all()

    # Get products for categorization
    dialysis_product = Product.query.filter(
        (Product.product_name == "Dialysis") | (Product.product_name == "HEMO DIALYSIS")
    ).first()

    # Build sales data structure:
    # sales[service_type][transaction_type][tender_name] = amount
    sales_data = {
        "Dialysis": {},  # Dialysis services
        "Diagnostics": {},  # All other services (Various Services, etc.)
    }

    # Get all transaction types
    transaction_types = (
        TransactionType.query.filter_by(active=True)
        .order_by(TransactionType.sort_order.desc())
        .all()
    )

    # Initialize structure
    for tt in transaction_types:
        sales_data["Dialysis"][tt.type_code] = {}
        sales_data["Diagnostics"][tt.type_code] = {}

    # Process transactions
    for txn in transactions:
        txn_type_code = (
            txn.transaction_type.type_code if txn.transaction_type else "walk_in"
        )

        # Determine if this transaction has dialysis services
        is_dialysis = False
        if dialysis_product and txn.transaction_details:
            is_dialysis = any(
                detail.product_id == dialysis_product.id
                for detail in txn.transaction_details
            )

        service_category = "Dialysis" if is_dialysis else "Diagnostics"

        # Add tender amounts
        for tender_record in txn.transaction_tenders:
            tender_name = (
                tender_record.tender.tender_name if tender_record.tender else "Unknown"
            )

            if txn_type_code not in sales_data[service_category]:
                sales_data[service_category][txn_type_code] = {}

            current_amount = sales_data[service_category][txn_type_code].get(
                tender_name, 0
            )
            sales_data[service_category][txn_type_code][tender_name] = (
                current_amount + tender_record.amount
            )

    # Calculate totals
    totals = {
        "Dialysis": {},
        "Diagnostics": {},
        "grand_total": 0,
    }

    for service_type in ["Dialysis", "Diagnostics"]:
        service_total = 0
        for txn_type_code in sales_data[service_type]:
            txn_type_total = sum(sales_data[service_type][txn_type_code].values())
            totals[service_type][txn_type_code] = txn_type_total
            service_total += txn_type_total
        totals[service_type]["total"] = service_total
        totals["grand_total"] += service_total

    # Get all tenders for consistent ordering
    all_tenders = Tender.query.order_by(Tender.sort_order.desc()).all()
    tender_order = {t.tender_name: t.sort_order for t in all_tenders}

    # Get unique tenders used in this report
    all_tender_names = set()
    for service_type in sales_data:
        for txn_type in sales_data[service_type]:
            all_tender_names.update(sales_data[service_type][txn_type].keys())

    # Sort tenders by their sort_order
    sorted_tenders = sorted(
        all_tender_names, key=lambda t: tender_order.get(t, 0), reverse=True
    )

    context = {
        "start_date": start_date,
        "end_date": end_date,
        "transaction_types": transaction_types,
        "sales_data": sales_data,
        "totals": totals,
        "sorted_tenders": sorted_tenders,
        "app_label": app_label,
        "transaction_count": len(transactions),
        "generated_at": datetime.now(),
    }

    return render_template("daily_sales/sales_summary_report.html", **context)


@bp.route("/sales_summary_excel", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def sales_summary_excel():
    """Export sales summary report to Excel"""
    import io

    from flask import send_file

    from ...register.product import Product

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        flash("openpyxl library not installed. Cannot export to Excel.", "danger")
        return redirect(url_for("daily_sales.sales_summary_report"))

    # Get date range from query params
    today = date.today()
    start_date_str = request.args.get(
        "start_date", str(date(today.year, today.month, 1))
    )
    end_date_str = request.args.get("end_date", str(today))

    try:
        start_date = date.fromisoformat(start_date_str)
        end_date = date.fromisoformat(end_date_str)
    except ValueError:
        start_date = date(today.year, today.month, 1)
        end_date = today

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    # Query transactions (same logic as report)
    transactions = Transaction.query.filter(
        Transaction.record_date >= str(start_date),
        Transaction.record_date <= str(end_date),
        Transaction.submitted.isnot(None),
        Transaction.cancelled.is_(None),
    ).all()

    dialysis_product = Product.query.filter(
        (Product.product_name == "Dialysis") | (Product.product_name == "HEMO DIALYSIS")
    ).first()

    sales_data = {"Dialysis": {}, "Diagnostics": {}}
    transaction_types = (
        TransactionType.query.filter_by(active=True)
        .order_by(TransactionType.sort_order.desc())
        .all()
    )

    for tt in transaction_types:
        sales_data["Dialysis"][tt.type_code] = {}
        sales_data["Diagnostics"][tt.type_code] = {}

    for txn in transactions:
        txn_type_code = (
            txn.transaction_type.type_code if txn.transaction_type else "walk_in"
        )
        is_dialysis = False
        if dialysis_product and txn.transaction_details:
            is_dialysis = any(
                detail.product_id == dialysis_product.id
                for detail in txn.transaction_details
            )

        service_category = "Dialysis" if is_dialysis else "Diagnostics"

        for tender_record in txn.transaction_tenders:
            tender_name = (
                tender_record.tender.tender_name if tender_record.tender else "Unknown"
            )
            if txn_type_code not in sales_data[service_category]:
                sales_data[service_category][txn_type_code] = {}
            current_amount = sales_data[service_category][txn_type_code].get(
                tender_name, 0
            )
            sales_data[service_category][txn_type_code][tender_name] = (
                current_amount + tender_record.amount
            )

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Summary"

    # Styles
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    section_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    section_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill(
        start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
    )
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title
    ws["A1"] = "Sales Summary Report"
    ws["A1"].font = title_font
    ws["A2"] = f"{start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}"
    ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    # Get tenders
    all_tenders = Tender.query.order_by(Tender.sort_order.desc()).all()
    tender_order = {t.tender_name: t.sort_order for t in all_tenders}
    all_tender_names = set()
    for service_type in sales_data:
        for txn_type in sales_data[service_type]:
            all_tender_names.update(sales_data[service_type][txn_type].keys())
    sorted_tenders = sorted(
        all_tender_names, key=lambda t: tender_order.get(t, 0), reverse=True
    )

    row = 5

    # For each service type
    for service_type in ["Diagnostics", "Dialysis"]:
        # Service type header
        ws.cell(row, 1, service_type).font = section_font
        ws.cell(row, 1).fill = section_fill
        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=len(sorted_tenders) + 2,
        )
        row += 1

        # Column headers
        ws.cell(row, 1, "Transaction Type").font = header_font
        for col_idx, tender in enumerate(sorted_tenders, start=2):
            ws.cell(row, col_idx, tender).font = header_font
        ws.cell(row, len(sorted_tenders) + 2, "Total").font = header_font
        row += 1

        # Transaction type rows
        service_total = 0
        for tt in transaction_types:
            if tt.type_code in sales_data[service_type]:
                ws.cell(row, 1, tt.type_name)
                row_total = 0
                for col_idx, tender in enumerate(sorted_tenders, start=2):
                    amount = sales_data[service_type][tt.type_code].get(tender, 0)
                    if amount > 0:
                        ws.cell(row, col_idx, amount).number_format = "#,##0.00"
                    row_total += amount
                ws.cell(
                    row, len(sorted_tenders) + 2, row_total
                ).number_format = "#,##0.00"
                ws.cell(row, len(sorted_tenders) + 2).font = Font(bold=True)
                service_total += row_total
                row += 1

        # Service total
        ws.cell(row, 1, f"{service_type} Total").font = Font(bold=True)
        ws.cell(row, 1).fill = total_fill
        ws.cell(row, len(sorted_tenders) + 2, service_total).number_format = "#,##0.00"
        ws.cell(row, len(sorted_tenders) + 2).font = Font(bold=True)
        ws.cell(row, len(sorted_tenders) + 2).fill = total_fill
        row += 2

    # Adjust column widths
    ws.column_dimensions["A"].width = 25
    for col_idx in range(2, len(sorted_tenders) + 3):
        ws.column_dimensions[chr(64 + col_idx)].width = 15

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"sales_summary_{start_date}_{end_date}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ---------------------------------------------------------------------------
# Internal Report class
# ---------------------------------------------------------------------------


class _Report:
    def __init__(self):
        self.report_date: date = ph_today()
        self.sales: dict = {}  # {type_code: {tender_name: amount}}

    @property
    def total_sales(self):
        return sum(sum(v.values()) for v in self.sales.values())

    @property
    def total_diagnostic_sales(self):
        return sum(
            sum(v.values()) for code, v in self.sales.items() if code != "dialysis"
        )

    @property
    def total_dialysis_sales(self):
        return sum(self.sales.get("dialysis", {}).values())


# ---------------------------------------------------------------------------
# Accountabilities Routes
# ---------------------------------------------------------------------------


@bp.route("/accountabilities", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def accountabilities():
    """Main accountabilities page"""
    # Get filter parameters
    start_date_str = request.args.get("start_date")
    end_date_str = request.args.get("end_date")
    category_id = request.args.get("category_id")

    # Default to current month if no dates provided (use Philippine timezone)
    today = ph_today()
    if not start_date_str:
        start_date = date(today.year, today.month, 1)
    else:
        start_date = date.fromisoformat(start_date_str)

    if not end_date_str:
        end_date = today
    else:
        end_date = date.fromisoformat(end_date_str)

    # Get all fund categories
    fund_categories = (
        FundCategory.query.filter_by(active=True)
        .order_by(FundCategory.sort_order)
        .all()
    )

    # Build query for funds received
    received_query = FundReceived.query.filter(
        FundReceived.record_date >= str(start_date),
        FundReceived.record_date <= str(end_date),
    )
    if category_id:
        received_query = received_query.filter_by(fund_category_id=int(category_id))
    funds_received = received_query.order_by(
        FundReceived.record_date.desc(), FundReceived.id.desc()
    ).all()

    # Build query for funds disbursed
    disbursed_query = FundDisbursed.query.filter(
        FundDisbursed.record_date >= str(start_date),
        FundDisbursed.record_date <= str(end_date),
    )
    if category_id:
        disbursed_query = disbursed_query.filter_by(fund_category_id=int(category_id))
    funds_disbursed = disbursed_query.order_by(
        FundDisbursed.record_date.desc(), FundDisbursed.id.desc()
    ).all()

    # Calculate totals
    total_received = sum(f.amount for f in funds_received if f.status == "posted")
    total_disbursed = sum(f.amount for f in funds_disbursed if f.status == "posted")

    context = {
        "fund_categories": fund_categories,
        "funds_received": funds_received,
        "funds_disbursed": funds_disbursed,
        "total_received": total_received,
        "total_disbursed": total_disbursed,
        "start_date": start_date,
        "end_date": end_date,
        "selected_category_id": int(category_id) if category_id else None,
        "app_label": app_label,
    }
    return render_template("daily_sales/accountabilities.html", **context)


@bp.route("/fund_received/new", methods=["GET", "POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def new_fund_received():
    """Create new fund received record"""
    if request.method == "POST":
        try:
            fund_category_id = request.form.get("fund_category_id")
            amount = float(request.form.get("amount", 0))
            record_date = request.form.get("record_date")
            reference_number = request.form.get("reference_number", "").strip()
            description = request.form.get("description", "").strip()

            if not fund_category_id or not record_date or amount <= 0:
                flash("Please provide category, date, and valid amount", "danger")
                return redirect(url_for("daily_sales.accountabilities"))

            fund_received = FundReceived(
                fund_category_id=int(fund_category_id),
                amount=amount,
                record_date=record_date,
                reference_number=reference_number,
                description=description,
                status="draft",  # Start as draft, requires approval
                created_by_id=current_user.id,
            )

            db.session.add(fund_received)
            db.session.flush()

            # Log fund received creation
            audit_log_create(
                module="fund_received",
                record_id=fund_received.id,
                record_identifier=f"Fund Received #{fund_received.id} - {fund_received.fund_category.category_name if fund_received.fund_category else 'N/A'} - ₱{fund_received.amount:,.2f}",
                new_values=model_to_dict(
                    fund_received,
                    [
                        "fund_category_id",
                        "amount",
                        "record_date",
                        "reference_number",
                        "description",
                        "status",
                    ],
                ),
                notes="Draft fund received created",
            )

            db.session.commit()

            flash("Fund received record created as draft", "success")
            return redirect(url_for("daily_sales.accountabilities"))

        except Exception as e:
            db.session.rollback()
            flash(f"Error creating fund received: {e!s}", "danger")
            return redirect(url_for("daily_sales.accountabilities"))

    # GET request - redirect to main page
    return redirect(url_for("daily_sales.accountabilities"))


@bp.route("/fund_disbursed/new", methods=["GET", "POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def new_fund_disbursed():
    """Create new fund disbursed record"""
    if request.method == "POST":
        try:
            fund_category_id = request.form.get("fund_category_id")
            amount = float(request.form.get("amount", 0))
            record_date = request.form.get("record_date")
            reference_number = request.form.get("reference_number", "").strip()
            description = request.form.get("description", "").strip()

            if not fund_category_id or not record_date or amount <= 0:
                flash("Please provide category, date, and valid amount", "danger")
                return redirect(url_for("daily_sales.accountabilities"))

            fund_disbursed = FundDisbursed(
                fund_category_id=int(fund_category_id),
                amount=amount,
                record_date=record_date,
                reference_number=reference_number,
                description=description,
                status="draft",  # Start as draft, requires approval
                created_by_id=current_user.id,
            )

            db.session.add(fund_disbursed)
            db.session.flush()

            # Log fund disbursed creation
            audit_log_create(
                module="fund_disbursed",
                record_id=fund_disbursed.id,
                record_identifier=f"Fund Disbursed #{fund_disbursed.id} - {fund_disbursed.fund_category.category_name if fund_disbursed.fund_category else 'N/A'} - ₱{fund_disbursed.amount:,.2f}",
                new_values=model_to_dict(
                    fund_disbursed,
                    [
                        "fund_category_id",
                        "amount",
                        "record_date",
                        "reference_number",
                        "description",
                        "status",
                    ],
                ),
                notes="Draft fund disbursed created",
            )

            db.session.commit()

            flash("Fund disbursed record created as draft", "success")
            return redirect(url_for("daily_sales.accountabilities"))

        except Exception as e:
            db.session.rollback()
            flash(f"Error creating fund disbursed: {e!s}", "danger")
            return redirect(url_for("daily_sales.accountabilities"))

    # GET request - redirect to main page
    return redirect(url_for("daily_sales.accountabilities"))


@bp.route("/fund_received/<int:id>/submit", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def submit_fund_received(id):
    """Submit fund received record for approval"""
    fund = FundReceived.query.get_or_404(id)

    if fund.status != "draft":
        flash("Only draft records can be submitted", "warning")
        return redirect(url_for("daily_sales.accountabilities"))

    fund.status = "submitted"
    fund.submitted_by_id = current_user.id
    fund.submitted_at = datetime.utcnow()

    # Log submission
    audit_log_status_change(
        module="fund_received",
        record_id=fund.id,
        record_identifier=f"Fund Received #{fund.id}",
        action="submitted",
        old_status="draft",
        new_status="submitted",
        notes=f"Submitted for approval - Amount: ₱{fund.amount:,.2f}",
    )

    db.session.commit()
    flash("Fund received submitted for approval", "success")
    return redirect(url_for("daily_sales.accountabilities"))


@bp.route("/fund_received/<int:id>/approve", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def approve_fund_received(id):
    """Approve fund received record"""
    fund = FundReceived.query.get_or_404(id)

    if fund.status != "submitted":
        flash("Only submitted records can be approved", "warning")
        return redirect(url_for("daily_sales.pending_approval"))

    fund.status = "posted"
    fund.approved_by_id = current_user.id
    fund.approved_at = datetime.utcnow()

    # Log approval
    audit_log_status_change(
        module="fund_received",
        record_id=fund.id,
        record_identifier=f"Fund Received #{fund.id}",
        action="approved",
        old_status="submitted",
        new_status="posted",
        notes=f"Approved and posted - Amount: ₱{fund.amount:,.2f}",
    )

    db.session.commit()
    flash("Fund received approved successfully", "success")
    return redirect(url_for("daily_sales.pending_approval"))


@bp.route("/fund_received/<int:id>/cancel", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def cancel_fund_received(id):
    """Request cancellation of fund received record"""
    fund = FundReceived.query.get_or_404(id)

    if fund.status == "cancelled":
        flash("Record is already cancelled", "warning")
        return redirect(url_for("daily_sales.accountabilities"))

    if fund.status == "pending_cancellation":
        flash("Cancellation request already pending approval", "warning")
        return redirect(url_for("daily_sales.accountabilities"))

    try:
        reason = request.form.get("reason", "").strip()

        # Admin/SuperUser can cancel immediately
        if current_user.admin or current_user.superuser:
            old_status = fund.status
            fund.status = "cancelled"
            fund.cancelled_by_id = current_user.id
            fund.cancelled_at = datetime.now()
            fund.cancellation_reason = reason if reason else "Cancelled by admin"

            # Log cancellation
            audit_log_status_change(
                module="fund_received",
                record_id=fund.id,
                record_identifier=f"Fund Received #{fund.id}",
                action="cancelled",
                old_status=old_status,
                new_status="cancelled",
                reason=reason if reason else "Cancelled by admin",
            )

            db.session.commit()
            flash("Fund received record cancelled successfully", "success")
        else:
            # Staff must request cancellation
            old_status = fund.status
            fund.status = "pending_cancellation"
            fund.cancelled_by_id = current_user.id  # Who requested
            fund.cancellation_reason = (
                reason if reason else "Cancellation requested by user"
            )

            # Log cancellation request
            audit_log_status_change(
                module="fund_received",
                record_id=fund.id,
                record_identifier=f"Fund Received #{fund.id}",
                action="cancellation_requested",
                old_status=old_status,
                new_status="pending_cancellation",
                reason=reason if reason else "Cancellation requested by user",
            )

            db.session.commit()
            flash("Cancellation request submitted. Waiting for admin approval.", "info")

    except Exception as e:
        db.session.rollback()
        flash(f"Error submitting cancellation request: {e!s}", "danger")
    return redirect(url_for("daily_sales.accountabilities"))


@bp.route("/fund_disbursed/<int:id>/submit", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def submit_fund_disbursed(id):
    """Submit fund disbursed record for approval"""
    fund = FundDisbursed.query.get_or_404(id)

    if fund.status != "draft":
        flash("Only draft records can be submitted", "warning")
        return redirect(url_for("daily_sales.accountabilities"))

    fund.status = "submitted"
    fund.submitted_by_id = current_user.id
    fund.submitted_at = datetime.utcnow()

    # Log submission
    audit_log_status_change(
        module="fund_disbursed",
        record_id=fund.id,
        record_identifier=f"Fund Disbursed #{fund.id}",
        action="submitted",
        old_status="draft",
        new_status="submitted",
        notes=f"Submitted for approval - Amount: ₱{fund.amount:,.2f}",
    )

    db.session.commit()
    flash("Fund disbursed submitted for approval", "success")
    return redirect(url_for("daily_sales.accountabilities"))


@bp.route("/fund_disbursed/<int:id>/approve", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def approve_fund_disbursed(id):
    """Approve fund disbursed record"""
    fund = FundDisbursed.query.get_or_404(id)

    if fund.status != "submitted":
        flash("Only submitted records can be approved", "warning")
        return redirect(url_for("daily_sales.pending_approval"))

    fund.status = "posted"
    fund.approved_by_id = current_user.id
    fund.approved_at = datetime.utcnow()

    # Audit log
    audit_log_status_change(
        module="fund_disbursed",
        record_id=fund.id,
        record_identifier=f"Fund Disbursed #{fund.id}",
        action="approved",
        old_status="submitted",
        new_status="posted",
        notes=f"Approved and posted - Amount: ₱{fund.amount:,.2f}",
    )

    db.session.commit()
    flash("Fund disbursed approved successfully", "success")
    return redirect(url_for("daily_sales.pending_approval"))


@bp.route("/fund_disbursed/<int:id>/cancel", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def cancel_fund_disbursed(id):
    """Request cancellation of fund disbursed record"""
    fund = FundDisbursed.query.get_or_404(id)

    if fund.status == "cancelled":
        flash("Record is already cancelled", "warning")
        return redirect(url_for("daily_sales.accountabilities"))

    if fund.status == "pending_cancellation":
        flash("Cancellation request already pending approval", "warning")
        return redirect(url_for("daily_sales.accountabilities"))

    try:
        reason = request.form.get("reason", "").strip()

        # Admin/SuperUser can cancel immediately
        if current_user.admin or current_user.superuser:
            old_status = fund.status
            fund.status = "cancelled"
            fund.cancelled_by_id = current_user.id
            fund.cancelled_at = datetime.now()
            fund.cancellation_reason = reason if reason else "Cancelled by admin"

            # Audit log
            audit_log_status_change(
                module="fund_disbursed",
                record_id=fund.id,
                record_identifier=f"Fund Disbursed #{fund.id}",
                action="cancelled",
                old_status=old_status,
                new_status="cancelled",
                reason=reason if reason else "Cancelled by admin",
            )

            db.session.commit()
            flash("Fund disbursed record cancelled successfully", "success")
        else:
            # Staff must request cancellation
            old_status = fund.status
            fund.status = "pending_cancellation"
            fund.cancelled_by_id = current_user.id  # Who requested
            fund.cancellation_reason = (
                reason if reason else "Cancellation requested by user"
            )

            # Audit log
            audit_log_status_change(
                module="fund_disbursed",
                record_id=fund.id,
                record_identifier=f"Fund Disbursed #{fund.id}",
                action="cancellation_requested",
                old_status=old_status,
                new_status="pending_cancellation",
                reason=reason if reason else "Cancellation requested by user",
            )

            db.session.commit()
            flash("Cancellation request submitted. Waiting for admin approval.", "info")

    except Exception as e:
        db.session.rollback()
        flash(f"Error submitting cancellation request: {e!s}", "danger")
    return redirect(url_for("daily_sales.accountabilities"))


@bp.route("/fund_received/<int:id>/approve_cancellation", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def approve_fund_received_cancellation(id):
    """Approve cancellation request for fund received (admin only)"""
    if not (current_user.admin or current_user.superuser):
        flash("Only admins can approve cancellations", "danger")
        return redirect(url_for("daily_sales.pending_fund_cancellations"))

    fund = FundReceived.query.get_or_404(id)

    if fund.status != "pending_cancellation":
        flash("No pending cancellation request for this record", "warning")
        return redirect(url_for("daily_sales.pending_fund_cancellations"))

    try:
        fund.status = "cancelled"
        fund.cancelled_at = datetime.now()
        # cancelled_by_id already set by requester, we track approval in notes
        if fund.cancellation_reason:
            fund.cancellation_reason += f" | Approved by {current_user.user_name}"
        else:
            fund.cancellation_reason = f"Approved by {current_user.user_name}"

        db.session.commit()
        flash("Cancellation request approved", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error approving cancellation: {e!s}", "danger")

    return redirect(url_for("daily_sales.pending_fund_cancellations"))


@bp.route("/fund_received/<int:id>/reject_cancellation", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def reject_fund_received_cancellation(id):
    """Reject cancellation request for fund received (admin only)"""
    if not (current_user.admin or current_user.superuser):
        flash("Only admins can reject cancellations", "danger")
        return redirect(url_for("daily_sales.pending_fund_cancellations"))

    fund = FundReceived.query.get_or_404(id)

    if fund.status != "pending_cancellation":
        flash("No pending cancellation request for this record", "warning")
        return redirect(url_for("daily_sales.pending_fund_cancellations"))

    try:
        # Revert to previous status (assume it was posted)
        fund.status = "posted"
        fund.cancelled_by_id = None
        fund.cancellation_reason = f"Cancellation rejected by {current_user.user_name}"

        db.session.commit()
        flash("Cancellation request rejected", "info")
    except Exception as e:
        db.session.rollback()
        flash(f"Error rejecting cancellation: {e!s}", "danger")

    return redirect(url_for("daily_sales.pending_fund_cancellations"))


@bp.route("/fund_disbursed/<int:id>/approve_cancellation", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def approve_fund_disbursed_cancellation(id):
    """Approve cancellation request for fund disbursed (admin only)"""
    if not (current_user.admin or current_user.superuser):
        flash("Only admins can approve cancellations", "danger")
        return redirect(url_for("daily_sales.pending_fund_cancellations"))

    fund = FundDisbursed.query.get_or_404(id)

    if fund.status != "pending_cancellation":
        flash("No pending cancellation request for this record", "warning")
        return redirect(url_for("daily_sales.pending_fund_cancellations"))

    try:
        fund.status = "cancelled"
        fund.cancelled_at = datetime.now()
        if fund.cancellation_reason:
            fund.cancellation_reason += f" | Approved by {current_user.user_name}"
        else:
            fund.cancellation_reason = f"Approved by {current_user.user_name}"

        db.session.commit()
        flash("Cancellation request approved", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error approving cancellation: {e!s}", "danger")

    return redirect(url_for("daily_sales.pending_fund_cancellations"))


@bp.route("/fund_disbursed/<int:id>/reject_cancellation", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def reject_fund_disbursed_cancellation(id):
    """Reject cancellation request for fund disbursed (admin only)"""
    if not (current_user.admin or current_user.superuser):
        flash("Only admins can reject cancellations", "danger")
        return redirect(url_for("daily_sales.pending_fund_cancellations"))

    fund = FundDisbursed.query.get_or_404(id)

    if fund.status != "pending_cancellation":
        flash("No pending cancellation request for this record", "warning")
        return redirect(url_for("daily_sales.pending_fund_cancellations"))

    try:
        # Revert to previous status (assume it was posted)
        fund.status = "posted"
        fund.cancelled_by_id = None
        fund.cancellation_reason = f"Cancellation rejected by {current_user.user_name}"

        db.session.commit()
        flash("Cancellation request rejected", "info")
    except Exception as e:
        db.session.rollback()
        flash(f"Error rejecting cancellation: {e!s}", "danger")

    return redirect(url_for("daily_sales.pending_fund_cancellations"))


@bp.route("/pending_fund_cancellations", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def pending_fund_cancellations():
    """View all pending cancellation requests (admin only)"""
    if not (current_user.admin or current_user.superuser):
        flash("Only admins can view pending cancellations", "danger")
        return redirect(url_for("daily_sales.accountabilities"))

    # Get all pending cancellation requests
    pending_received = (
        FundReceived.query.filter_by(status="pending_cancellation")
        .order_by(FundReceived.record_date.desc())
        .all()
    )
    pending_disbursed = (
        FundDisbursed.query.filter_by(status="pending_cancellation")
        .order_by(FundDisbursed.record_date.desc())
        .all()
    )

    context = {
        "app_label": app_label,
        "pending_received": pending_received,
        "pending_disbursed": pending_disbursed,
        "total_pending": len(pending_received) + len(pending_disbursed),
    }
    return render_template("daily_sales/pending_fund_cancellations.html", **context)


# ---------------------------------------------------------------------------
# Petty Cash Management Routes
# ---------------------------------------------------------------------------


@bp.route("/petty_cash_management", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def petty_cash_management():
    """Petty Cash Management page"""
    from application.extensions import next_control_number

    # Get filter parameters
    start_date_str = request.args.get("start_date")
    end_date_str = request.args.get("end_date")

    today = ph_today()
    if not start_date_str:
        start_date = date(today.year, today.month, 1)
    else:
        start_date = date.fromisoformat(start_date_str)

    if not end_date_str:
        end_date = today
    else:
        end_date = date.fromisoformat(end_date_str)

    # Get vouchers within date range
    vouchers = (
        PettyCashVoucher.query.filter(
            PettyCashVoucher.record_date >= str(start_date),
            PettyCashVoucher.record_date <= str(end_date),
        )
        .order_by(PettyCashVoucher.record_date.desc(), PettyCashVoucher.id.desc())
        .all()
    )

    # Get reimbursement reports
    reports = ReimbursementReport.query.order_by(
        ReimbursementReport.created_date.desc()
    ).all()

    # Get pending reports (for reimbursement linking)
    pending_reports = ReimbursementReport.query.filter_by(status="pending").all()

    # Get reimbursement received records within date range
    reimbursements_received = (
        ReimbursementReceived.query.filter(
            ReimbursementReceived.record_date >= str(start_date),
            ReimbursementReceived.record_date <= str(end_date),
        )
        .order_by(
            ReimbursementReceived.record_date.desc(), ReimbursementReceived.id.desc()
        )
        .all()
    )

    # Generate next numbers
    next_pcv = (
        next_control_number(PettyCashVoucher, "pcv_number")
        if PettyCashVoucher.query.count() > 0
        else "PCV-0001"
    )
    next_ref = (
        next_control_number(ReimbursementReceived, "reference_number")
        if ReimbursementReceived.query.count() > 0
        else "REF-0001"
    )

    # Get all payees for autocomplete
    payees = Payee.query.filter_by(active=True).order_by(Payee.name).all()

    # Get all bank accounts for autocomplete
    bank_accounts = (
        BankAccount.query.filter_by(active=True)
        .order_by(BankAccount.account_name)
        .all()
    )

    # Calculate Petty Cash Balance
    # Formula: (Funds Received - Funds Disbursed) for Petty Cash only - Submitted PCV (not yet reimbursed)

    # Get Petty Cash category only (exclude Change Fund)
    petty_cash_category = FundCategory.query.filter(
        FundCategory.category_name == "Petty Cash"
    ).first()
    category_id = petty_cash_category.id if petty_cash_category else None

    # 1. Get all Funds Received (posted/submitted) for Petty Cash only up to today
    total_funds_received = (
        sum(
            f.amount
            for f in FundReceived.query.filter(
                FundReceived.record_date <= str(today),
                FundReceived.status.in_(["posted", "submitted"]),
                FundReceived.fund_category_id == category_id,
            ).all()
        )
        if category_id
        else 0
    )

    # 2. Get all Funds Disbursed (posted/submitted) for Petty Cash only up to today
    total_funds_disbursed = (
        sum(
            f.amount
            for f in FundDisbursed.query.filter(
                FundDisbursed.record_date <= str(today),
                FundDisbursed.status.in_(["posted", "submitted"]),
                FundDisbursed.fund_category_id == category_id,
            ).all()
        )
        if category_id
        else 0
    )

    # 3. Get submitted/posted PCV that have not been reimbursed
    # (PCVs that are submitted or posted but not yet in a reimbursement report)
    unreimbursed_pcv = sum(
        pcv.amount
        for pcv in PettyCashVoucher.query.filter(
            PettyCashVoucher.status.in_(["submitted", "posted"]),
            PettyCashVoucher.record_date <= str(today),
        ).all()
    )

    petty_cash_balance = total_funds_received - total_funds_disbursed - unreimbursed_pcv

    context = {
        "app_label": app_label,
        "vouchers": vouchers,
        "reports": reports,
        "pending_reports": pending_reports,
        "reimbursements_received": reimbursements_received,
        "next_pcv": next_pcv,
        "next_ref": next_ref,
        "payees": payees,
        "bank_accounts": bank_accounts,
        "start_date": str(start_date),
        "end_date": str(end_date),
        "today": str(today),
        "petty_cash_balance": petty_cash_balance,
        "total_funds_received": total_funds_received,
        "total_funds_disbursed": total_funds_disbursed,
        "unreimbursed_pcv": unreimbursed_pcv,
    }
    return render_template("daily_sales/petty_cash_management.html", **context)


@bp.route("/petty_cash/voucher/new", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def new_petty_cash_voucher():
    """Create new petty cash voucher"""
    try:
        record_date = request.form.get("date")
        pcv_number = request.form.get("pcv_number", "").strip()
        payee_id = request.form.get("payee_id", "").strip()
        amount = float(request.form.get("amount", 0))
        description = request.form.get("description", "").strip()

        if not record_date or not pcv_number or not payee_id or amount <= 0:
            flash("Please provide date, PCV number, payee, and valid amount", "danger")
            return redirect(url_for("daily_sales.petty_cash_management"))

        # Check if PCV number already exists
        existing_pcv = PettyCashVoucher.query.filter_by(pcv_number=pcv_number).first()
        if existing_pcv:
            flash(
                f"PCV number {pcv_number} already exists. Please use a different number.",
                "danger",
            )
            return redirect(url_for("daily_sales.petty_cash_management"))

        # Verify payee exists
        payee = Payee.query.get(int(payee_id))
        if not payee:
            flash("Invalid payee selected", "danger")
            return redirect(url_for("daily_sales.petty_cash_management"))

        # Create voucher
        voucher = PettyCashVoucher(
            pcv_number=pcv_number,
            record_date=record_date,
            payee_id=payee.id,
            amount=amount,
            description=description,
            status="draft",
            created_by_id=current_user.id,
        )

        db.session.add(voucher)
        db.session.flush()  # Ensure ID is assigned

        # Log creation with audit trail
        from .audit_logger import log_create

        log_create(
            record_type="petty_cash_voucher",
            instance=voucher,
            user=current_user,
            notes="New petty cash voucher created",
        )

        db.session.commit()

        flash(f"Petty Cash Voucher {pcv_number} created successfully!", "success")
        return redirect(url_for("daily_sales.petty_cash_management"))

    except Exception as e:
        db.session.rollback()
        flash(f"Error creating voucher: {e!s}", "danger")
        return redirect(url_for("daily_sales.petty_cash_management"))


@bp.route("/petty_cash/voucher/<int:voucher_id>/data", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def get_voucher_data(voucher_id):
    """Get voucher data as JSON for editing"""
    voucher = PettyCashVoucher.query.get_or_404(voucher_id)

    return jsonify(
        {
            "id": voucher.id,
            "record_date": voucher.record_date,
            "pcv_number": voucher.pcv_number,
            "payee_id": voucher.payee_id,
            "payee_name": voucher.payee.name if voucher.payee else "",
            "amount": float(voucher.amount),
            "description": voucher.description or "",
            "created_by_name": str(voucher.created_by) if voucher.created_by else "N/A",
            "status": voucher.status,
        }
    )


@bp.route("/petty_cash/voucher/<int:voucher_id>/edit", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def edit_petty_cash_voucher(voucher_id):
    """Edit draft petty cash voucher with audit trail"""
    voucher = PettyCashVoucher.query.get_or_404(voucher_id)

    if voucher.status != "draft":
        flash("Only draft vouchers can be edited", "danger")
        return redirect(url_for("daily_sales.petty_cash_management"))

    try:
        # Import audit logger
        from .audit_logger import get_model_snapshot, log_update

        # Capture old snapshot before making changes
        old_snapshot = get_model_snapshot(voucher)

        record_date = request.form.get("date")
        pcv_number = request.form.get("pcv_number", "").strip()
        payee_id = request.form.get("payee_id", "").strip()
        amount = float(request.form.get("amount", 0))
        description = request.form.get("description", "").strip()

        if not record_date or not pcv_number or not payee_id or amount <= 0:
            flash("Please provide date, PCV number, payee, and valid amount", "danger")
            return redirect(url_for("daily_sales.petty_cash_management"))

        # Check if PCV number already exists (excluding current voucher)
        existing_pcv = PettyCashVoucher.query.filter(
            PettyCashVoucher.pcv_number == pcv_number, PettyCashVoucher.id != voucher_id
        ).first()
        if existing_pcv:
            flash(
                f"PCV number {pcv_number} already exists. Please use a different number.",
                "danger",
            )
            return redirect(url_for("daily_sales.petty_cash_management"))

        # Verify payee exists
        payee = Payee.query.get(int(payee_id))
        if not payee:
            flash("Invalid payee selected", "danger")
            return redirect(url_for("daily_sales.petty_cash_management"))

        # Update voucher
        voucher.record_date = record_date
        voucher.pcv_number = pcv_number
        voucher.payee_id = payee.id
        voucher.amount = amount
        voucher.description = description

        # Log the update with audit trail
        log_update(
            record_type="petty_cash_voucher",
            instance=voucher,
            old_snapshot=old_snapshot,
            user=current_user,
            notes="Voucher edited while in draft status",
        )

        db.session.commit()
        flash(f"Petty Cash Voucher {pcv_number} updated successfully!", "success")
        return redirect(url_for("daily_sales.petty_cash_management"))

    except Exception as e:
        db.session.rollback()
        flash(f"Error updating voucher: {e!s}", "danger")
        return redirect(url_for("daily_sales.petty_cash_management"))


@bp.route("/petty_cash/voucher/<int:voucher_id>/submit", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def submit_petty_cash_voucher(voucher_id):
    """Submit draft voucher"""
    voucher = PettyCashVoucher.query.get_or_404(voucher_id)

    if not voucher.can_submit:
        flash("Only draft vouchers can be submitted", "warning")
        return redirect(url_for("daily_sales.petty_cash_management"))

    voucher.status = "submitted"
    voucher.submitted_by_id = current_user.id
    voucher.submitted_at = datetime.now()

    # Log submission with audit trail
    from .audit_logger import log_status_change

    log_status_change(
        record_type="petty_cash_voucher",
        instance=voucher,
        action="submitted",
        user=current_user,
        notes="Voucher submitted for approval",
    )

    db.session.commit()
    flash(f"Voucher {voucher.pcv_number} submitted successfully!", "success")
    return redirect(url_for("daily_sales.petty_cash_management"))


@bp.route("/petty_cash/voucher/<int:voucher_id>/post", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def post_petty_cash_voucher(voucher_id):
    """Post submitted voucher (admin only)"""
    # Check where to redirect after approval
    next_page = (
        request.form.get("next") or request.args.get("next") or "petty_cash_management"
    )

    if not current_user.admin and not current_user.superuser:
        flash("Only admins can post vouchers", "danger")
        return redirect(url_for(f"daily_sales.{next_page}"))

    voucher = PettyCashVoucher.query.get_or_404(voucher_id)

    if not voucher.can_post:
        flash("Only submitted vouchers can be posted", "warning")
        return redirect(url_for(f"daily_sales.{next_page}"))

    voucher.status = "posted"
    voucher.posted_by_id = current_user.id
    voucher.posted_at = datetime.now()

    # Log approval with audit trail
    from .audit_logger import log_status_change

    log_status_change(
        record_type="petty_cash_voucher",
        instance=voucher,
        action="approved",
        user=current_user,
        notes="Voucher posted/approved by admin",
    )

    db.session.commit()
    flash(f"Voucher {voucher.pcv_number} posted successfully!", "success")
    return redirect(url_for(f"daily_sales.{next_page}"))


@bp.route("/petty_cash/voucher/<int:voucher_id>/cancel", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def cancel_petty_cash_voucher(voucher_id):
    """Cancel voucher"""
    voucher = PettyCashVoucher.query.get_or_404(voucher_id)

    if not voucher.can_cancel:
        flash("This voucher cannot be cancelled", "warning")
        return redirect(url_for("daily_sales.petty_cash_management"))

    reason = request.form.get("reason", "").strip()

    voucher.status = "cancelled"
    voucher.cancelled_by_id = current_user.id
    voucher.cancelled_at = datetime.now()
    voucher.cancelled_reason = reason

    # Log cancellation with audit trail
    from .audit_logger import log_status_change

    log_status_change(
        record_type="petty_cash_voucher",
        instance=voucher,
        action="cancelled",
        user=current_user,
        reason=reason,
        notes="Voucher cancelled by user",
    )

    db.session.commit()
    flash(f"Voucher {voucher.pcv_number} cancelled", "info")
    return redirect(url_for("daily_sales.petty_cash_management"))


@bp.route("/petty_cash/voucher/<int:voucher_id>/delete", methods=["POST", "GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def petty_cash_voucher_delete(voucher_id):
    """Delete draft petty cash voucher"""
    voucher = PettyCashVoucher.query.get_or_404(voucher_id)

    # Only allow deleting draft vouchers
    if voucher.status != "draft":
        flash("Only draft vouchers can be deleted", "danger")
        return redirect(url_for("daily_sales.petty_cash_management"))

    try:
        pcv_number = voucher.pcv_number

        # Log deletion with audit trail before deleting
        from .audit_logger import log_delete

        log_delete(
            record_type="petty_cash_voucher",
            instance=voucher,
            user=current_user,
            reason="User deleted draft voucher",
            notes=f"Voucher {pcv_number} deleted by user",
        )

        db.session.delete(voucher)
        db.session.commit()
        flash(f"Voucher {pcv_number} deleted successfully", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting voucher: {e!s}", "danger")

    return redirect(url_for("daily_sales.petty_cash_management"))


@bp.route("/petty_cash/voucher/<int:voucher_id>/history", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def get_voucher_history(voucher_id):
    """Get audit history for a voucher"""
    from .audit_logger import get_audit_history

    voucher = PettyCashVoucher.query.get_or_404(voucher_id)

    # Get audit history
    audit_logs = get_audit_history("petty_cash_voucher", voucher_id)

    # Format audit logs for JSON response
    history = []
    for log in audit_logs:
        history.append(
            {
                "action": log.action,
                "user": log.user.user_name if log.user else "System",
                "user_full_name": str(log.user) if log.user else "System",
                "timestamp": log.formatted_timestamp
                if hasattr(log, "formatted_timestamp")
                else str(log.created_at),
                "summary": log.summary
                if hasattr(log, "summary")
                else f"{log.action} by {log.user.user_name if log.user else 'System'}",
                "old_values": log.old_values,
                "new_values": log.new_values,
                "changed_fields": log.changed_fields,
                "reason": log.reason,
                "notes": log.notes,
            }
        )

    return jsonify({"voucher_number": voucher.pcv_number, "history": history})


@bp.route("/petty_cash/reimbursement_report/create", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def create_reimbursement_report():
    """Create reimbursement report from selected vouchers"""
    from application.extensions import next_control_number

    try:
        voucher_ids = request.form.getlist("voucher_ids[]")

        if not voucher_ids:
            flash("No vouchers selected", "warning")
            return redirect(url_for("daily_sales.petty_cash_management"))

        # Get selected vouchers
        vouchers = PettyCashVoucher.query.filter(
            PettyCashVoucher.id.in_(voucher_ids), PettyCashVoucher.status == "posted"
        ).all()

        if not vouchers:
            flash("No valid posted vouchers found", "warning")
            return redirect(url_for("daily_sales.petty_cash_management"))

        # Calculate period
        dates = [datetime.strptime(v.record_date, "%Y-%m-%d").date() for v in vouchers]
        period_start = str(min(dates))
        period_end = str(max(dates))

        # Generate report number
        report_number = (
            next_control_number(ReimbursementReport, "report_number")
            if ReimbursementReport.query.count() > 0
            else "RPT-001"
        )

        # Create report
        report = ReimbursementReport(
            report_number=report_number,
            created_date=str(ph_today()),
            period_start=period_start,
            period_end=period_end,
            status="pending",
            prepared_by_id=current_user.id,
        )

        db.session.add(report)
        db.session.flush()

        # Link vouchers to report
        total = 0
        for voucher in vouchers:
            voucher.status = "for_reimbursement"
            voucher.reimbursement_report_id = report.id
            total += voucher.amount

        report.total_amount = total

        db.session.commit()
        flash(
            f"Reimbursement Report {report_number} created with {len(vouchers)} vouchers (₱{total:,.2f})",
            "success",
        )
        return redirect(url_for("daily_sales.petty_cash_management"))

    except Exception as e:
        db.session.rollback()
        flash(f"Error creating report: {e!s}", "danger")
        return redirect(url_for("daily_sales.petty_cash_management"))


@bp.route("/petty_cash/reimbursement/new", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def new_reimbursement_received():
    """Record reimbursement received"""
    try:
        record_date = request.form.get("date")
        reference_number = request.form.get("reference_number", "").strip()
        bank_account_id = request.form.get("bank_account_id", "").strip()
        amount = float(request.form.get("amount", 0))
        notes = request.form.get("notes", "").strip()
        report_ids = request.form.getlist("report_ids[]")

        if (
            not record_date
            or not reference_number
            or not bank_account_id
            or amount <= 0
        ):
            flash(
                "Please provide date, reference number, bank account, and valid amount",
                "danger",
            )
            return redirect(url_for("daily_sales.petty_cash_management"))

        # Check if reference number already exists
        existing_ref = ReimbursementReceived.query.filter_by(
            reference_number=reference_number
        ).first()
        if existing_ref:
            flash(
                f"Reference number {reference_number} already exists. Please use a different number.",
                "danger",
            )
            return redirect(url_for("daily_sales.petty_cash_management"))

        # Verify bank account exists
        bank_account = BankAccount.query.get(int(bank_account_id))
        if not bank_account:
            flash("Invalid bank account selected", "danger")
            return redirect(url_for("daily_sales.petty_cash_management"))

        # Create reimbursement received record
        reimbursement = ReimbursementReceived(
            reference_number=reference_number,
            record_date=record_date,
            bank_account_id=bank_account.id,
            amount=amount,
            notes=notes,
            status="draft",
            created_by_id=current_user.id,
        )

        db.session.add(reimbursement)
        db.session.flush()

        # Update linked reports status to reimbursed
        if report_ids:
            reports = ReimbursementReport.query.filter(
                ReimbursementReport.id.in_(report_ids)
            ).all()
            for report in reports:
                report.status = "reimbursed"
                # Update vouchers under this report
                for voucher in report.vouchers:
                    voucher.status = "reimbursed"

        db.session.commit()
        flash(f"Reimbursement {reference_number} recorded successfully!", "success")
        return redirect(url_for("daily_sales.petty_cash_management"))

    except Exception as e:
        db.session.rollback()
        flash(f"Error recording reimbursement: {e!s}", "danger")
        return redirect(url_for("daily_sales.petty_cash_management"))


@bp.route("/reimbursement_received/<int:id>/submit", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def submit_reimbursement_received(id):
    """Submit reimbursement received for approval"""
    reimbursement = ReimbursementReceived.query.get_or_404(id)

    if reimbursement.status != "draft":
        flash("Only draft reimbursements can be submitted", "warning")
        return redirect(url_for("daily_sales.petty_cash_management"))

    reimbursement.status = "submitted"
    reimbursement.submitted_by_id = current_user.id
    reimbursement.submitted_at = datetime.utcnow()
    db.session.commit()

    flash(
        f"Reimbursement {reimbursement.reference_number} submitted for approval",
        "success",
    )
    return redirect(url_for("daily_sales.petty_cash_management"))


@bp.route("/reimbursement_received/<int:id>/approve", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def approve_reimbursement_received(id):
    """Approve reimbursement received"""
    reimbursement = ReimbursementReceived.query.get_or_404(id)

    if reimbursement.status != "submitted":
        flash("Only submitted reimbursements can be approved", "warning")
        return redirect(url_for("daily_sales.petty_cash_management"))

    reimbursement.status = "posted"
    reimbursement.approved_by_id = current_user.id
    reimbursement.approved_at = datetime.utcnow()
    db.session.commit()

    flash(
        f"Reimbursement {reimbursement.reference_number} approved successfully",
        "success",
    )
    return redirect(url_for("daily_sales.petty_cash_management"))


@bp.route("/reimbursement_received/<int:id>/delete", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def delete_reimbursement_received(id):
    """Delete reimbursement received"""
    reimbursement = ReimbursementReceived.query.get_or_404(id)

    try:
        ref_number = reimbursement.reference_number
        db.session.delete(reimbursement)
        db.session.commit()
        flash(f"Reimbursement {ref_number} deleted successfully", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting reimbursement: {e!s}", "danger")

    return redirect(url_for("daily_sales.petty_cash_management"))


@bp.route("/petty_cash/payees", methods=["GET"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def get_payees():
    """API endpoint for payee autocomplete"""
    payees = Payee.query.filter_by(active=True).order_by(Payee.name).all()
    return {"payees": [p.name for p in payees]}


@bp.route("/petty_cash/payee/add", methods=["POST"])
@login_required
@roles_accepted([ROLES_ACCEPTED])
def add_payee():
    """Add new payee"""
    name = request.form.get("name", "").strip()

    if not name:
        return {"success": False, "message": "Payee name is required"}

    # Check if exists
    existing = Payee.query.filter_by(name=name).first()
    if existing:
        return {"success": False, "message": "Payee already exists"}

    payee = Payee(name=name, active=True)
    db.session.add(payee)
    db.session.commit()

    return {"success": True, "payee": name}


# ---------------------------------------------------------------------------
# Change Request Routes
# ---------------------------------------------------------------------------

from .change_request_views import (
    change_history,
    change_requests_list,
    request_transaction_cancellation,
    request_transaction_change,
    review_change_request,
    transaction_history,
)

bp.route("/transaction/<int:transaction_id>/request_change", methods=["GET", "POST"])(
    login_required(roles_accepted([ROLES_ACCEPTED])(request_transaction_change))
)

bp.route(
    "/transaction/<int:transaction_id>/request_cancellation", methods=["GET", "POST"]
)(login_required(roles_accepted([ROLES_ACCEPTED])(request_transaction_cancellation)))

bp.route("/change_requests", methods=["GET"])(
    login_required(roles_accepted([ROLES_ACCEPTED])(change_requests_list))
)

bp.route("/change_request/<int:request_id>/review", methods=["POST"])(
    login_required(roles_accepted([ROLES_ACCEPTED])(review_change_request))
)

bp.route("/change_history", methods=["GET"])(
    login_required(roles_accepted([ROLES_ACCEPTED])(change_history))
)

bp.route("/transaction/<int:transaction_id>/history", methods=["GET"])(
    login_required(roles_accepted([ROLES_ACCEPTED])(transaction_history))
)
